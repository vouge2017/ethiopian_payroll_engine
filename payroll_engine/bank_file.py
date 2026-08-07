"""
Bank File Generator — Ethiopian Payroll Engine

Generates bulk payment files for Ethiopian banks and mobile wallets.
Supports: CBE, Dashen, Awash, Telebirr.

Key rules from the implementation blueprint:
- Account numbers must be TEXT (not numbers) to prevent Excel scientific notation
- No commas in numbers — force 2 decimal places, string data type
- CBE traditional: 13 numeric digits (starts with 1000...)
- Telebirr/CBE Birr/mobile wallets: 9 digits starting with 9 or 7
- Pre-validation catches bad account numbers before file generation
"""

import csv
import io
import re
from typing import Any

# Account validation patterns
ACCOUNT_PATTERNS = {
    'cbe': {
        'name': 'Commercial Bank of Ethiopia (CBE)',
        'pattern': r'^1\d{12}$',
        'description': '13 digits starting with 1 (e.g., 1000123456789)',
        'example': '1000123456789',
    },
    'dashen': {
        'name': 'Dashen Bank',
        'pattern': r'^\d{13}$',
        'description': '13 numeric digits',
        'example': '0990123456789',
    },
    'awash': {
        'name': 'Awash Bank',
        'pattern': r'^\d{13}$',
        'description': '13 numeric digits',
        'example': '0990123456789',
    },
    'boa': {
        'name': 'Bank of Abyssinia',
        'pattern': r'^\d{13}$',
        'description': '13 numeric digits',
        'example': '0990123456789',
    },
    'wegagen': {
        'name': 'Wegagen Bank',
        'pattern': r'^\d{13}$',
        'description': '13 numeric digits',
        'example': '0990123456789',
    },
    'nib': {
        'name': 'NIB International Bank',
        'pattern': r'^\d{13}$',
        'description': '13 numeric digits',
        'example': '0990123456789',
    },
    'bunna': {
        'name': 'Bunna Bank',
        'pattern': r'^\d{13}$',
        'description': '13 numeric digits',
        'example': '0990123456789',
    },
    'zemen': {
        'name': 'Zemen Bank',
        'pattern': r'^\d{13}$',
        'description': '13 numeric digits',
        'example': '0990123456789',
    },
    'lion': {
        'name': 'Lion International Bank',
        'pattern': r'^\d{13}$',
        'description': '13 numeric digits',
        'example': '0990123456789',
    },
    'telebirr': {
        'name': 'Telebirr / Mobile Wallet',
        'pattern': r'^(0?9|0?7)\d{8}$',
        'description': '09 or 07 followed by 8 digits',
        'example': '912345678',
    },
    'mpesa': {
        'name': 'M-Pesa (Safaricom)',
        'pattern': r'^(0?7)\d{8}$',
        'description': '9 digits starting with 7 (or 10 with leading 0)',
        'example': '712345678',
    },
}


def validate_account_number(account: str, bank: str) -> tuple:
    """
    Validate an account number against the bank's format rules.

    Returns:
        (is_valid: bool, error_message: str or None)
    """
    if not account or not account.strip():
        return False, 'Account number is empty'

    account = account.strip()

    # Detect bank from account if not specified
    if bank not in ACCOUNT_PATTERNS:
        # Try to auto-detect
        for key, info in ACCOUNT_PATTERNS.items():
            if re.match(info['pattern'], account):
                bank = key
                break
        else:
            return False, f"Unknown bank format: '{bank}'. Supported: {', '.join(ACCOUNT_PATTERNS.keys())}"

    pattern = ACCOUNT_PATTERNS[bank]['pattern']
    if not re.match(pattern, account):
        expected = ACCOUNT_PATTERNS[bank]['description']
        return False, f"Invalid {ACCOUNT_PATTERNS[bank]['name']} account: '{account}'. Expected: {expected}"

    return True, None


# Narrative template presets
NARRATIVE_TEMPLATES = {
    'id_name': '{period} salary - {id} {name}',
    'name_only': '{period} salary - {name}',
    'id_only': '{period} salary - {id}',
    'period_name': '{period} - {name}',
    'custom': None,  # User provides their own
}


def get_narrative(template_key: str, emp_id: str, name: str, period: str, custom_template: str | None = None) -> str:
    """
    Generate narrative text from a template.

    Args:
        template_key: One of NARRATIVE_TEMPLATES keys, or 'custom'
        emp_id: Employee ID
        name: Employee name
        period: Pay period (e.g., 'July 2025')
        custom_template: Custom template string (used when template_key='custom')
            Available placeholders: {period}, {id}, {name}

    Returns:
        Formatted narrative string
    """
    if template_key == 'custom' and custom_template:
        template = custom_template
    else:
        template = NARRATIVE_TEMPLATES.get(template_key, NARRATIVE_TEMPLATES['id_name'])

    return template.format(period=period, id=emp_id, name=name)


def format_amount(amount: float, decimals: int = 2) -> str:
    """
    Format amount for bank file: no commas, forced decimal places, string type.

    Args:
        amount: The amount to format
        decimals: Number of decimal places (default 2, some banks want 0)

    Bad:  "12,500.50" or 12500.5
    Good: "12500.50"
    """
    return f'{amount:.{decimals}f}'


def validate_payroll_for_bank(
    employees_data: list[dict[str, Any]], bank: str = 'cbe', previous_payslips: dict[str, dict] | None = None
) -> list[dict[str, Any]]:
    """
    Pre-validate all employees before generating bank file.

    Catches:
    - Missing or invalid account numbers
    - Duplicate employees (same ID twice in same run)
    - Duplicate account numbers (same bank account for different employees)
    - Account number changes from previous run (suspicious)
    - Negative or zero net pay

    Args:
        employees_data: List of employee dicts
        bank: Bank key (cbe, dashen, awash, telebirr)
        previous_payslips: Dict mapping employee_id to previous payslip data
            (used to detect account changes)

    Returns:
        List of error/warning dicts (empty if all valid)
    """
    errors = []

    # --- Track duplicates within this run ---
    seen_ids = {}  # employee_id -> first occurrence index
    seen_accounts = {}  # account_number -> first employee_id

    for i, emp in enumerate(employees_data):
        emp_id = emp.get('id', '')
        emp_name = emp.get('name', '')

        # --- CHECK 1: Duplicate employee ID ---
        if emp_id in seen_ids:
            errors.append(
                {
                    'employee_id': emp_id,
                    'name': emp_name,
                    'field': 'employee_id',
                    'error': f'DUPLICATE: Employee {emp_id} appears twice in this run '
                    f'(first at row {seen_ids[emp_id] + 1}, again at row {i + 1})',
                    'severity': 'BLOCK',
                }
            )
            continue  # Skip further checks for duplicate entry
        seen_ids[emp_id] = i

        # --- CHECK 2: Missing account number ---
        account = emp.get('bank', '').strip()
        if not account:
            errors.append(
                {
                    'employee_id': emp_id,
                    'name': emp_name,
                    'field': 'bank_or_telebirr',
                    'error': 'Missing bank/Telebirr account number',
                    'severity': 'BLOCK',
                }
            )
            continue

        # Extract account number from format
        if ':' in account:
            parts = account.split(':', 1)
            account_type = parts[0].lower()
            account_number = parts[1].strip()
        else:
            account_number = account
            account_type = bank

        # --- CHECK 3: Invalid account format ---
        bank_key = account_type
        if account_type == 'bank':
            bank_key = bank
        is_valid, error_msg = validate_account_number(account_number, bank_key)
        if not is_valid:
            errors.append(
                {
                    'employee_id': emp_id,
                    'name': emp_name,
                    'field': 'bank_or_telebirr',
                    'error': error_msg,
                    'severity': 'BLOCK',
                }
            )

        # --- CHECK 4: Same bank account used by different employees ---
        if account_number in seen_accounts:
            errors.append(
                {
                    'employee_id': emp_id,
                    'name': emp_name,
                    'field': 'bank_or_telebirr',
                    'error': f'DUPLICATE ACCOUNT: Bank account {account_number} '
                    f'is also assigned to employee {seen_accounts[account_number]}. '
                    f'One account cannot receive two salaries.',
                    'severity': 'BLOCK',
                }
            )
        else:
            seen_accounts[account_number] = emp_id

        # --- CHECK 5: Account number changed from previous run ---
        if previous_payslips and emp_id in previous_payslips:
            prev = previous_payslips[emp_id]
            prev_account = prev.get('bank', '').strip()
            if prev_account and ':' in prev_account:
                prev_account = prev_account.split(':', 1)[1].strip()
            if prev_account and account_number != prev_account:
                errors.append(
                    {
                        'employee_id': emp_id,
                        'name': emp_name,
                        'field': 'bank_or_telebirr',
                        'error': f'ACCOUNT CHANGED: Was {prev_account} last month, '
                        f'now {account_number}. Verify this is correct.',
                        'severity': 'FLAG',  # Not a block, but needs confirmation
                    }
                )

        # --- CHECK 6: Negative or zero net pay ---
        net = emp.get('net', 0)
        if net <= 0:
            errors.append(
                {
                    'employee_id': emp_id,
                    'name': emp_name,
                    'field': 'net_pay',
                    'error': f'Net pay must be positive, got {net}',
                    'severity': 'BLOCK',
                }
            )

    return errors


def generate_csv(
    employees_data: list[dict[str, Any]],
    bank: str = 'cbe',
    company_name: str = '',
    period: str = '',
    narrative_template: str = 'id_name',
    custom_narrative: str | None = None,
    decimals: int = 2,
) -> bytes:
    """
    Generate a bank-ready CSV file for bulk payment upload.

    Args:
        employees_data: List of employee dicts with id, name, bank, net
        bank: Bank key (cbe, dashen, awash, telebirr)
        company_name: Company name for the header
        period: Pay period (e.g., "July 2025")

    Returns:
        CSV file as bytes
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Header row — matches CBE bulk upload format
    writer.writerow(['account_number', 'amount', 'narrative', 'currency'])

    # Data rows
    for emp in employees_data:
        account = emp.get('bank', '').strip()

        # Extract account number from format
        if ':' in account:
            account = account.split(':', 1)[1].strip()

        # Format as TEXT — no commas, configurable decimal places
        amount = format_amount(emp.get('net', 0), decimals=decimals)

        # Narrative from template
        narrative = get_narrative(
            narrative_template,
            emp_id=emp.get('id', ''),
            name=emp.get('name', 'Unknown'),
            period=period,
            custom_template=custom_narrative,
        )

        # CSV injection prevention
        from payroll_engine.security import prevent_csv_injection

        writer.writerow(
            [
                account,
                amount,
                prevent_csv_injection(narrative),
                'ETB',
            ]
        )

    return output.getvalue().encode('utf-8')


def generate_xlsx(
    employees_data: list[dict[str, Any]],
    bank: str = 'cbe',
    company_name: str = '',
    period: str = '',
    narrative_template: str = 'id_name',
    custom_narrative: str | None = None,
    decimals: int = 2,
) -> bytes:
    """
    Generate a bank-ready Excel file with account numbers as TEXT
    (prevents Excel scientific notation on 13-digit numbers).

    Args:
        employees_data: List of employee dicts
        bank: Bank key
        company_name: Company name
        period: Pay period

    Returns:
        XLSX file as bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        # Fallback to CSV
        return generate_csv(employees_data, bank, company_name, period)

    from datetime import date as _date

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Bank Transfer'

    # Styles
    title_font = Font(bold=True, size=16, color='1A5276')
    subtitle_font = Font(bold=True, size=12, color='333333')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_align = Alignment(horizontal='right', vertical='center')
    name_align = Alignment(horizontal='left', vertical='center')
    totals_font = Font(bold=True, size=11, color='1A5276')
    totals_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    text_format = '@'  # Explicit text format — prevents scientific notation
    etb_format = '#,##0.00'

    # --- Title block ---
    ws.merge_cells('A1:D1')
    ws['A1'] = company_name
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('A2:D2')
    ws['A2'] = f'Bank Transfer File — {period}' if period else 'Bank Transfer File'
    ws['A2'].font = subtitle_font

    ws.merge_cells('A3:D3')
    ws['A3'] = f'Generated: {_date.today().strftime("%d %B %Y")}'
    ws['A3'].font = Font(italic=True, color='666666')

    # --- Headers (row 5) ---
    headers = ['Account Number', 'Amount (ETB)', 'Narrative', 'Currency']
    col_widths = [20, 16, 40, 10]
    for col, (header, width) in enumerate(zip(headers, col_widths, strict=False), 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # --- Data rows ---
    total_amount = 0
    for i, emp in enumerate(employees_data, 1):
        account = emp.get('bank', '').strip()
        if ':' in account:
            account = account.split(':', 1)[1].strip()

        amount = emp.get('net', 0)
        total_amount += amount
        name = emp.get('name', 'Unknown')
        emp_id = emp.get('id', '')
        narrative = get_narrative(
            narrative_template,
            emp_id=emp_id,
            name=name,
            period=period,
            custom_template=custom_narrative,
        )

        row = 5 + i

        # Account number: force TEXT format (prevents Excel scientific notation)
        cell = ws.cell(row=row, column=1, value=account)
        cell.number_format = text_format
        cell.alignment = name_align
        cell.border = thin_border

        # Amount
        cell = ws.cell(row=row, column=2, value=amount)
        cell.number_format = etb_format
        cell.alignment = data_align
        cell.border = thin_border

        ws.cell(row=row, column=3, value=narrative).alignment = name_align
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4, value='ETB').alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4).border = thin_border

    # --- Totals row ---
    totals_row = 5 + len(employees_data) + 1
    cell = ws.cell(row=totals_row, column=1, value='TOTAL')
    cell.font = totals_font
    cell.fill = totals_fill
    cell.border = thin_border
    cell = ws.cell(row=totals_row, column=2, value=total_amount)
    cell.font = totals_font
    cell.fill = totals_fill
    cell.number_format = etb_format
    cell.alignment = data_align
    cell.border = thin_border
    ws.cell(row=totals_row, column=3).fill = totals_fill
    ws.cell(row=totals_row, column=3).border = thin_border
    ws.cell(row=totals_row, column=4).fill = totals_fill
    ws.cell(row=totals_row, column=4).border = thin_border

    # --- Print setup ---
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
