"""
Excel Import Utility

Reads .xlsx files and returns data as list of dicts.
Supports common Ethiopian payroll spreadsheet formats.

Handles:
- Column header normalization (case-insensitive, strips spaces)
- Phone number normalization (Ethiopian format)
- Salary parsing (removes commas, ETB prefix)
- Empty row skipping
- Multiple sheet support
"""

import re
from decimal import Decimal, InvalidOperation


def read_xlsx(file_path_or_bytes, sheet_name=None):
    """
    Read an Excel file and return rows as list of dicts.

    Args:
        file_path_or_bytes: file path (str) or file-like object
        sheet_name: specific sheet to read (default: first sheet)

    Returns:
        list of dicts with normalized column names as keys
    """
    import openpyxl

    if isinstance(file_path_or_bytes, (str, bytes)):
        wb = openpyxl.load_workbook(file_path_or_bytes, read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(file_path_or_bytes, read_only=True, data_only=True)

    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Normalize headers
    headers = []
    for h in rows[0]:
        if h is None:
            headers.append('')
        else:
            # Normalize: lowercase, strip, replace spaces with underscores
            h = str(h).strip().lower()
            h = re.sub(r'[^a-z0-9_]', '_', h)
            h = re.sub(r'_+', '_', h).strip('_')
            headers.append(h)

    # Map common header variations
    header_map = {
        'emp_id': 'employee_id',
        'emp_code': 'employee_id',
        'staff_id': 'employee_id',
        'id': 'employee_id',
        'full_name': 'name',
        'employee_name': 'name',
        'first_name': 'name',
        'fname': 'name',
        'mobile': 'phone',
        'tel': 'phone',
        'telephone': 'phone',
        'mobile_number': 'phone',
        'phone_number': 'phone',
        'salary': 'basic_salary',
        'basic': 'basic_salary',
        'base_salary': 'basic_salary',
        'monthly_salary': 'basic_salary',
        'allowance': 'allowances',
        'total_allowance': 'allowances',
        'dept': 'department',
        'dep': 'department',
        'pos': 'position',
        'job_title': 'position',
        'designation': 'position',
        'bank': 'bank_account',
        'account': 'bank_account',
        'bank_account_number': 'bank_account',
        'account_no': 'bank_account',
        'tin_number': 'tin',
        'tax_id': 'tin',
        'tax_identification_number': 'tin',
    }

    normalized_headers = []
    for h in headers:
        normalized_headers.append(header_map.get(h, h))

    # Read data rows
    result = []
    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue

        row_dict = {}
        for i, val in enumerate(row):
            if i < len(normalized_headers) and normalized_headers[i]:
                key = normalized_headers[i]
                if val is not None:
                    row_dict[key] = str(val).strip() if not isinstance(val, (int, float)) else val
                else:
                    row_dict[key] = ''

        if any(v for v in row_dict.values()):
            result.append(row_dict)

    wb.close()
    return result


def parse_salary(value):
    """Parse salary value from various formats."""
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    if not value:
        return Decimal('0')

    # Remove common prefixes/suffixes
    value = str(value).strip()
    value = value.replace('ETB', '').replace('etb', '').replace('Birr', '').replace('birr', '')
    value = value.replace(',', '').replace(' ', '').strip()

    try:
        return Decimal(value)
    except (InvalidOperation, ValueError):
        return Decimal('0')


def normalize_phone(phone_str):
    """Normalize Ethiopian phone number."""
    if not phone_str:
        return None

    from payroll_engine.models import validate_ethiopian_phone

    is_valid, normalized, _error = validate_ethiopian_phone(str(phone_str))
    return normalized if is_valid else None


def detect_file_type(filename):
    """Detect if file is CSV or Excel based on extension."""
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    if ext in ('xlsx', 'xls'):
        return 'excel'
    elif ext == 'csv':
        return 'csv'
    return 'unknown'


def read_file(file_storage):
    """
    Read uploaded file (CSV or Excel) and return rows as list of dicts.

    Args:
        file_storage: Flask FileStorage object

    Returns:
        list of dicts
    """
    filename = file_storage.filename
    file_type = detect_file_type(filename)

    if file_type == 'excel':
        return read_xlsx(file_storage.stream)
    elif file_type == 'csv':
        import csv
        import io

        content = file_storage.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        return [dict(row) for row in reader]
    else:
        raise ValueError(f'Unsupported file type: {filename}')
