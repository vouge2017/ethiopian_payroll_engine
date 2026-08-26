"""
Excel-Compatible Payroll Engine — Deterministic, Explainable, Auditable

This module provides a complete Excel-based payroll workflow:
1. Import employee data from .xlsx/.csv
2. Calculate payroll with full deterministic audit trail
3. Generate multi-sheet Excel workbook with:
   - Payroll summary
   - Per-employee calculation flow (explainable)
   - Tax bracket breakdown
   - Exception report
   - Bank-ready payment file
   - Change detection vs previous period
   - Approval workflow sheet
4. Export bank files for Ethiopian banks

Design principles:
- Same inputs → same outputs (deterministic Decimal math)
- Every number has a formula trail
- Every exception has a severity and resolution path
- Approval is a state machine, not a checkbox
"""

import io
import json
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from typing import Any

Q = Decimal('0.01')


def _D(value) -> Decimal:
    """Safely convert any numeric type to Decimal."""
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal('0')


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass
class CalculationStep:
    """One step in the payroll calculation trail."""

    step_number: int
    label: str
    formula: str
    inputs: dict
    result: Decimal
    note: str = ''
    is_deduction: bool = False
    legal_reference: str = ''


@dataclass
class EmployeePayrollResult:
    """Complete payroll result for one employee with full audit trail."""

    # Identity
    employee_id: str
    employee_name: str
    department: str = ''
    position: str = ''

    # Inputs
    basic_salary: Decimal = Decimal('0')
    allowances: Decimal = Decimal('0')
    overtime_hours: Decimal = Decimal('0')
    overtime_type: str = 'day'

    # Allowance breakdown
    allowance_details: list = field(default_factory=list)
    exempt_allowances: Decimal = Decimal('0')
    taxable_allowances: Decimal = Decimal('0')

    # Overtime
    overtime_pay: Decimal = Decimal('0')
    overtime_rate: Decimal = Decimal('0')
    overtime_total_hours: Decimal = Decimal('0')

    # Calculation results
    gross: Decimal = Decimal('0')
    pension_employee: Decimal = Decimal('0')
    pension_employer: Decimal = Decimal('0')
    taxable: Decimal = Decimal('0')
    tax: Decimal = Decimal('0')
    net: Decimal = Decimal('0')

    # Sick leave
    sick_leave_reduction: Decimal = Decimal('0')

    # Deductions
    total_deductions: Decimal = Decimal('0')
    deduction_details: list = field(default_factory=list)

    # Calculation trail
    steps: list = field(default_factory=list)

    # Tax explanation
    tax_explanation: str = ''
    tax_breakdown: dict = field(default_factory=dict)

    # Effective rates
    effective_tax_rate: Decimal = Decimal('0')
    pension_tax_savings: Decimal = Decimal('0')

    # Exceptions
    exceptions: list = field(default_factory=list)

    # Bank info
    bank_account: str = ''
    bank_name: str = ''
    tin: str = ''

    # Status
    status: str = 'calculated'  # calculated, exception, approved, rejected
    exception_reason: str = ''


@dataclass
class ExceptionItem:
    """An exception found during payroll processing."""

    employee_id: str
    employee_name: str
    rule_code: str
    severity: str  # BLOCK, FLAG, WARN
    message: str
    hint: str = ''
    details: dict = field(default_factory=dict)
    overridden: bool = False
    override_reason: str = ''
    overridden_by: str = ''


@dataclass
class PayrollRunResult:
    """Complete result of a payroll run."""

    # Run metadata
    run_id: str = ''
    period: str = ''
    run_date: str = ''
    company_name: str = ''

    # Employee results
    employees: list = field(default_factory=list)

    # Totals
    total_gross: Decimal = Decimal('0')
    total_tax: Decimal = Decimal('0')
    total_pension_employee: Decimal = Decimal('0')
    total_pension_employer: Decimal = Decimal('0')
    total_net: Decimal = Decimal('0')
    total_overtime: Decimal = Decimal('0')
    total_deductions: Decimal = Decimal('0')
    employee_count: int = 0

    # Exceptions
    exceptions: list = field(default_factory=list)
    block_count: int = 0
    flag_count: int = 0
    warn_count: int = 0

    # Change detection
    previous_total_gross: Decimal = Decimal('0')
    previous_total_net: Decimal = Decimal('0')
    previous_employee_count: int = 0
    gross_delta: Decimal = Decimal('0')
    gross_delta_pct: Decimal = Decimal('0')
    new_hires: list = field(default_factory=list)
    departures: list = field(default_factory=list)
    salary_changes: list = field(default_factory=list)

    # Approval
    status: str = 'draft'  # draft, review, approved, locked
    approved_by: str = ''
    approved_at: str = ''
    approval_notes: str = ''

    # Bank file
    bank_file_data: bytes = b''
    bank_file_name: str = ''

    # Determinism proof
    calculation_hash: str = ''
    input_hash: str = ''


# ---------------------------------------------------------------------------
# Core calculation engine
# ---------------------------------------------------------------------------


class ExcelPayrollEngine:
    """
    Deterministic payroll calculation engine with Excel I/O.

    Usage:
        engine = ExcelPayrollEngine()
        result = engine.run_from_excel('payroll_input.xlsx', company_id=1)
        engine.export_to_excel(result, 'payroll_output.xlsx')
    """

    def __init__(self, company_id: int | None = None, for_date: str | date | None = None):
        self.company_id = company_id
        self.for_date = for_date
        self._tax_rule = None
        self._load_rules()

    def _load_rules(self):
        """Load tax rules from database or use defaults."""
        try:
            from payroll_engine.models import TaxRule

            self._tax_rule = TaxRule.get_active_rule(self.for_date)
        except Exception:
            self._tax_rule = None

    @property
    def tax_brackets(self):
        """Get tax brackets."""
        if self._tax_rule and self._tax_rule.brackets:
            return self._tax_rule.brackets
        return [
            {'min': 0, 'max': 2000, 'rate': 0},
            {'min': 2000, 'max': 4000, 'rate': 0.15},
            {'min': 4000, 'max': 7000, 'rate': 0.20},
            {'min': 7000, 'max': 10000, 'rate': 0.25},
            {'min': 10000, 'max': 14000, 'rate': 0.30},
            {'min': 14000, 'max': None, 'rate': 0.35},
        ]

    @property
    def pension_rates(self):
        """Get pension rates."""
        if self._tax_rule:
            return {
                'employee': self._tax_rule.pension_employee_rate,
                'employer': self._tax_rule.pension_employer_rate,
                'ceiling': self._tax_rule.pension_ceiling,
            }
        return {'employee': Decimal('0.07'), 'employer': Decimal('0.11'), 'ceiling': None}

    # ------------------------------------------------------------------
    # Calculation
    # ------------------------------------------------------------------

    def calculate_employee(
        self,
        employee_id: str,
        name: str,
        basic_salary: Decimal,
        allowances: Decimal = Decimal('0'),
        department: str = '',
        position: str = '',
        bank_account: str = '',
        bank_name: str = '',
        tin: str = '',
        overtime_hours: Decimal = Decimal('0'),
        overtime_type: str = 'day',
        allowance_records: list | None = None,
        deductions: list | None = None,
        sick_leave_reduction: Decimal = Decimal('0'),
    ) -> EmployeePayrollResult:
        """
        Calculate payroll for a single employee with full audit trail.

        Returns EmployeePayrollResult with every step documented.
        """
        from payroll_engine.overtime import calculate_total_overtime
        from payroll_engine.tax import calculate_tax, calculate_tax_breakdown, explain_tax_amharic

        # Validate inputs
        if _D(basic_salary) < 0:
            raise ValueError(f'basic_salary cannot be negative: {basic_salary}')
        if _D(allowances) < 0:
            raise ValueError(f'allowances cannot be negative: {allowances}')

        result = EmployeePayrollResult(
            employee_id=employee_id,
            employee_name=name,
            department=department,
            position=position,
            basic_salary=_D(basic_salary),
            allowances=_D(allowances),
            bank_account=bank_account,
            bank_name=bank_name,
            tin=tin,
        )

        steps = []
        step_num = 0

        # Step 1: Basic salary
        step_num += 1
        steps.append(
            CalculationStep(
                step_number=step_num,
                label='Basic Salary',
                formula='Input',
                inputs={'basic_salary': result.basic_salary},
                result=result.basic_salary,
                note='Monthly basic salary from contract',
            )
        )

        # Step 2: Allowances
        exempt_total = Decimal('0')
        taxable_total = Decimal('0')
        allowance_details = []

        if allowance_records:
            for rec in allowance_records:
                if not rec.get('is_active', True):
                    continue
                amt = _D(rec.get('amount', 0))
                treatment = rec.get('tax_treatment', 'taxable')
                if treatment == 'exempt':
                    exempt_total += amt
                elif treatment == 'partial':
                    cap = _D(rec.get('exempt_cap_amount', 0))
                    exempt_part = min(amt, cap) if cap > 0 else Decimal('0')
                    exempt_total += exempt_part
                    taxable_total += amt - exempt_part
                else:
                    taxable_total += amt
                # Per-allowance exempt/taxable split
                if treatment == 'exempt':
                    a_exempt = amt
                    a_taxable = Decimal('0')
                elif treatment == 'partial':
                    cap = _D(rec.get('exempt_cap_amount', 0))
                    a_exempt = min(amt, cap) if cap > 0 else Decimal('0')
                    a_taxable = amt - a_exempt
                else:
                    a_exempt = Decimal('0')
                    a_taxable = amt
                allowance_details.append(
                    {
                        'type': rec.get('allowance_type', 'other'),
                        'amount': amt,
                        'exempt': a_exempt,
                        'taxable': a_taxable,
                        'treatment': treatment,
                    }
                )
            total_allowances = exempt_total + taxable_total
        else:
            total_allowances = _D(allowances)
            taxable_total = total_allowances
            exempt_total = Decimal('0')

        result.exempt_allowances = exempt_total
        result.taxable_allowances = taxable_total
        result.allowance_details = allowance_details

        step_num += 1
        steps.append(
            CalculationStep(
                step_number=step_num,
                label='Total Allowances',
                formula='Sum of all allowances',
                inputs={'exempt': exempt_total, 'taxable': taxable_total},
                result=total_allowances,
                note=f'Exempt: ETB {exempt_total:,.2f}, Taxable: ETB {taxable_total:,.2f}',
            )
        )

        # Step 3: Overtime
        overtime_pay = Decimal('0')
        overtime_rate = Decimal('0')
        overtime_total_hours = Decimal('0')
        if _D(overtime_hours) > 0:
            ot_result = calculate_total_overtime(
                result.basic_salary,
                [{'hours': float(overtime_hours), 'type': overtime_type}],
            )
            overtime_pay = ot_result['total_pay']
            overtime_total_hours = ot_result['total_hours']
            if ot_result.get('entries'):
                overtime_rate = ot_result['entries'][0].get('hourly_rate', Decimal('0'))

        result.overtime_pay = overtime_pay
        result.overtime_rate = overtime_rate
        result.overtime_total_hours = overtime_total_hours

        if overtime_pay > 0:
            step_num += 1
            steps.append(
                CalculationStep(
                    step_number=step_num,
                    label='Overtime Pay',
                    formula=f'Hourly rate × hours × multiplier ({overtime_type})',
                    inputs={
                        'basic_salary': result.basic_salary,
                        'hours': overtime_total_hours,
                        'type': overtime_type,
                    },
                    result=overtime_pay,
                    note=f'Hourly rate: ETB {overtime_rate:,.2f}',
                    legal_reference='Proclamation No. 1156/2019, Articles 67-68',
                )
            )

        # Step 4: Gross
        gross = result.basic_salary + total_allowances + overtime_pay
        result.gross = gross.quantize(Q, rounding=ROUND_HALF_UP)

        step_num += 1
        steps.append(
            CalculationStep(
                step_number=step_num,
                label='Gross Salary',
                formula='Basic + Allowances + Overtime',
                inputs={
                    'basic': result.basic_salary,
                    'allowances': total_allowances,
                    'overtime': overtime_pay,
                },
                result=result.gross,
            )
        )

        # Step 5: Pension (BEFORE tax — legal requirement)
        rates = self.pension_rates
        emp_rate = Decimal(str(rates['employee']))
        empr_rate = Decimal(str(rates['employer']))
        ceiling = rates.get('ceiling')

        insurable = min(result.basic_salary, _D(ceiling)) if ceiling else result.basic_salary
        pension_emp = (insurable * emp_rate).quantize(Q, rounding=ROUND_HALF_UP)
        pension_empr = (insurable * empr_rate).quantize(Q, rounding=ROUND_HALF_UP)

        result.pension_employee = pension_emp
        result.pension_employer = pension_empr

        step_num += 1
        steps.append(
            CalculationStep(
                step_number=step_num,
                label='Employee Pension (7%)',
                formula=f'{emp_rate*100:.0f}% of basic salary',
                inputs={'basic_salary': result.basic_salary, 'rate': emp_rate, 'ceiling': ceiling},
                result=pension_emp,
                note='Deducted BEFORE tax (legal requirement)',
                is_deduction=True,
                legal_reference='Proclamation No. 1268/2022, Article 10',
            )
        )

        # Step 6: Taxable income
        taxable = gross - pension_emp - exempt_total
        taxable = max(Decimal('0'), taxable)
        result.taxable = taxable.quantize(Q, rounding=ROUND_HALF_UP)

        step_num += 1
        steps.append(
            CalculationStep(
                step_number=step_num,
                label='Taxable Income',
                formula='Gross − Pension − Exempt Allowances',
                inputs={
                    'gross': result.gross,
                    'pension': pension_emp,
                    'exempt_allowances': exempt_total,
                },
                result=result.taxable,
                note='This is the amount tax is calculated on',
            )
        )

        # Step 7: Tax
        tax = calculate_tax(taxable, self.for_date)
        result.tax = tax
        result.tax_explanation = explain_tax_amharic(taxable, self.for_date)
        result.tax_breakdown = calculate_tax_breakdown(taxable, self.for_date)

        step_num += 1
        steps.append(
            CalculationStep(
                step_number=step_num,
                label='Income Tax',
                formula='Progressive brackets on taxable income',
                inputs={'taxable_income': result.taxable},
                result=tax,
                note='Proclamation No. 1395/2025, Article 11',
                is_deduction=True,
                legal_reference='Proclamation No. 1395/2025, Article 11',
            )
        )

        # Step 8: Pension tax savings
        from payroll_engine.tax import calculate_tax as calc_tax

        tax_without_pension = calc_tax(gross - exempt_total, self.for_date)
        pension_savings = tax_without_pension - tax
        result.pension_tax_savings = pension_savings

        if pension_savings > 0:
            step_num += 1
            steps.append(
                CalculationStep(
                    step_number=step_num,
                    label='Pension Tax Savings',
                    formula='Tax without pension deduction − Tax with pension deduction',
                    inputs={
                        'tax_without_pension': tax_without_pension,
                        'tax_with_pension': tax,
                    },
                    result=pension_savings,
                    note=f'You save ETB {pension_savings:,.2f}/month because pension is deducted before tax',
                )
            )

        # Step 9: Net before deductions
        net_before = gross - tax - pension_emp
        step_num += 1
        steps.append(
            CalculationStep(
                step_number=step_num,
                label='Net Before Deductions',
                formula='Gross − Tax − Pension',
                inputs={'gross': result.gross, 'tax': tax, 'pension': pension_emp},
                result=net_before.quantize(Q, rounding=ROUND_HALF_UP),
            )
        )

        # Step 10: Sick leave reduction
        sick_reduction = _D(sick_leave_reduction)
        result.sick_leave_reduction = sick_reduction
        if sick_reduction > 0:
            step_num += 1
            steps.append(
                CalculationStep(
                    step_number=step_num,
                    label='Sick Leave Reduction',
                    formula='Pay reduction for days exceeding 30-day full-pay tier',
                    inputs={'reduction': sick_reduction},
                    result=sick_reduction,
                    note='Proclamation No. 1156/2019, Article 86',
                    is_deduction=True,
                )
            )

        # Step 11: Post-tax deductions
        total_deductions = Decimal('0')
        deduction_details = []
        if deductions:
            for ded in deductions:
                if not ded.get('is_active', True):
                    continue
                ded_amount = _D(ded.get('amount', 0))
                if ded_amount > 0:
                    total_deductions += ded_amount
                    deduction_details.append(
                        {
                            'type': ded.get('deduction_type', 'other'),
                            'label': ded.get('label', ''),
                            'amount': ded_amount,
                        }
                    )

        result.total_deductions = total_deductions
        result.deduction_details = deduction_details

        if total_deductions > 0:
            step_num += 1
            steps.append(
                CalculationStep(
                    step_number=step_num,
                    label='Post-Tax Deductions',
                    formula='Cost-sharing, court orders, loans, etc.',
                    inputs={'deductions': total_deductions},
                    result=total_deductions,
                    note=f'{len(deduction_details)} deduction(s)',
                    is_deduction=True,
                )
            )

        # Step 12: Final net
        net = net_before - sick_reduction - total_deductions
        net = max(Decimal('0'), net)
        result.net = net.quantize(Q, rounding=ROUND_HALF_UP)

        step_num += 1
        steps.append(
            CalculationStep(
                step_number=step_num,
                label='Net Pay',
                formula='Net Before Deductions − Sick Leave − Post-Tax Deductions',
                inputs={
                    'net_before': net_before.quantize(Q, rounding=ROUND_HALF_UP),
                    'sick_reduction': sick_reduction,
                    'post_tax_deductions': total_deductions,
                },
                result=result.net,
                note='What the employee takes home',
            )
        )

        # Effective tax rate
        if result.gross > 0:
            result.effective_tax_rate = (tax / result.gross * 100).quantize(Q, rounding=ROUND_HALF_UP)

        result.steps = steps
        return result

    # ------------------------------------------------------------------
    # Batch processing
    # ------------------------------------------------------------------

    def run_from_data(
        self,
        employees_data: list[dict],
        company_name: str = '',
        period: str = '',
        previous_payslips: dict | None = None,
    ) -> PayrollRunResult:
        """
        Process a batch of employees and return complete payroll results.

        Args:
            employees_data: List of dicts with employee data
            company_name: Company name for reports
            period: Pay period string
            previous_payslips: Previous period data for change detection

        Returns:
            PayrollRunResult with all employees, exceptions, and bank file
        """
        import hashlib

        run = PayrollRunResult(
            company_name=company_name,
            period=period,
            run_date=date.today().isoformat(),
        )

        # Determinism: hash the input data
        input_str = json.dumps(employees_data, sort_keys=True, default=str)
        run.input_hash = hashlib.sha256(input_str.encode()).hexdigest()[:16]

        # Process each employee
        for emp_data in employees_data:
            try:
                emp_result = self.calculate_employee(
                    employee_id=str(emp_data.get('employee_id', emp_data.get('id', ''))),
                    name=str(emp_data.get('name', '')),
                    basic_salary=_D(emp_data.get('basic_salary', emp_data.get('basic', 0))),
                    allowances=_D(emp_data.get('allowances', 0)),
                    department=str(emp_data.get('department', '')),
                    position=str(emp_data.get('position', '')),
                    bank_account=str(emp_data.get('bank_account', '')),
                    bank_name=str(emp_data.get('bank_name', emp_data.get('bank', ''))),
                    tin=str(emp_data.get('tin', '')),
                    overtime_hours=_D(emp_data.get('overtime_hours', 0)),
                    overtime_type=str(emp_data.get('overtime_type', 'day')),
                    allowance_records=emp_data.get('allowance_records'),
                    deductions=emp_data.get('deductions'),
                    sick_leave_reduction=_D(emp_data.get('sick_leave_reduction', 0)),
                )
                run.employees.append(emp_result)
            except Exception as e:
                # Capture calculation errors as exceptions
                exc = ExceptionItem(
                    employee_id=str(emp_data.get('employee_id', emp_data.get('id', ''))),
                    employee_name=str(emp_data.get('name', '')),
                    rule_code='CALCULATION_ERROR',
                    severity='BLOCK',
                    message=f'Calculation failed: {str(e)}',
                    hint='Check the input values for this employee.',
                )
                run.exceptions.append(exc)

        # Run validation
        exceptions = self._run_validation(run.employees, previous_payslips)
        run.exceptions.extend(exceptions)

        # Attach exceptions to individual employees
        for exc in run.exceptions:
            for emp in run.employees:
                if emp.employee_id == exc.employee_id:
                    emp.exceptions.append(exc)
                    if exc.severity == 'BLOCK':
                        emp.status = 'exception'
                        emp.exception_reason = exc.message

        # Compute totals
        run.employee_count = len(run.employees)
        run.total_gross = sum((e.gross for e in run.employees), Decimal('0'))
        run.total_tax = sum((e.tax for e in run.employees), Decimal('0'))
        run.total_pension_employee = sum((e.pension_employee for e in run.employees), Decimal('0'))
        run.total_pension_employer = sum((e.pension_employer for e in run.employees), Decimal('0'))
        run.total_net = sum((e.net for e in run.employees), Decimal('0'))
        run.total_overtime = sum((e.overtime_pay for e in run.employees), Decimal('0'))
        run.total_deductions = sum((e.total_deductions for e in run.employees), Decimal('0'))

        # Exception counts
        run.block_count = sum(1 for e in run.exceptions if e.severity == 'BLOCK' and not e.overridden)
        run.flag_count = sum(1 for e in run.exceptions if e.severity == 'FLAG' and not e.overridden)
        run.warn_count = sum(1 for e in run.exceptions if e.severity == 'WARN')

        # Change detection
        if previous_payslips:
            self._compute_changes(run, previous_payslips)

        # Determinism: hash the output
        output_str = json.dumps(
            {
                'employees': [
                    {
                        'id': e.employee_id,
                        'gross': str(e.gross),
                        'tax': str(e.tax),
                        'pension': str(e.pension_employee),
                        'net': str(e.net),
                    }
                    for e in run.employees
                ],
                'totals': {
                    'gross': str(run.total_gross),
                    'tax': str(run.total_tax),
                    'net': str(run.total_net),
                },
            },
            sort_keys=True,
        )
        run.calculation_hash = hashlib.sha256(output_str.encode()).hexdigest()[:16]

        # Generate bank file
        if run.block_count == 0:
            self._generate_bank_file(run)

        return run

    def run_from_excel(self, filepath: str, **kwargs) -> PayrollRunResult:
        """Import from Excel/CSV and process."""
        import csv
        import os

        ext = os.path.splitext(filepath)[1].lower()

        if ext == '.csv':
            # Handle CSV files directly
            with open(filepath, newline='', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                rows = [dict(row) for row in reader]
        else:
            from payroll_engine.excel_import import read_xlsx
            rows = read_xlsx(filepath)

        if not rows:
            raise ValueError('File is empty or has no data')

        # Normalize column names
        normalized = []
        for row in rows:
            emp = {}
            for key, val in row.items():
                k = str(key).strip().lower().replace(' ', '_')
                emp[k] = val
            normalized.append(emp)

        return self.run_from_data(normalized, **kwargs)

    # ------------------------------------------------------------------
    # Validation
    # ------------------------------------------------------------------

    def _run_validation(
        self,
        employees: list[EmployeePayrollResult],
        previous_payslips: dict | None,
    ) -> list[ExceptionItem]:
        """Run validation checks on calculated results."""
        exceptions = []

        # Check for duplicate IDs
        seen_ids = {}
        for emp in employees:
            if emp.employee_id in seen_ids:
                exceptions.append(
                    ExceptionItem(
                        employee_id=emp.employee_id,
                        employee_name=emp.employee_name,
                        rule_code='DUPLICATE_ID',
                        severity='BLOCK',
                        message=f'Duplicate employee ID: {emp.employee_id}',
                        hint='Each employee must have a unique ID.',
                    )
                )
            seen_ids[emp.employee_id] = emp.employee_name

        # Check for negative net pay
        for emp in employees:
            if emp.net < 0:
                exceptions.append(
                    ExceptionItem(
                        employee_id=emp.employee_id,
                        employee_name=emp.employee_name,
                        rule_code='NEGATIVE_NET',
                        severity='BLOCK',
                        message=f'Negative net pay: ETB {emp.net:,.2f}',
                        hint='Deductions exceed gross salary.',
                    )
                )

        # Check for missing bank account
        for emp in employees:
            if not emp.bank_account.strip():
                exceptions.append(
                    ExceptionItem(
                        employee_id=emp.employee_id,
                        employee_name=emp.employee_name,
                        rule_code='MISSING_BANK',
                        severity='BLOCK',
                        message='No bank account or Telebirr number',
                        hint='Required for salary disbursement.',
                    )
                )

        # Check pension calculation
        for emp in employees:
            if emp.basic_salary > 0:
                expected = (emp.basic_salary * Decimal('0.07')).quantize(Q, rounding=ROUND_HALF_UP)
                if abs(emp.pension_employee - expected) > Decimal('0.01'):
                    exceptions.append(
                        ExceptionItem(
                            employee_id=emp.employee_id,
                            employee_name=emp.employee_name,
                            rule_code='PENSION_MISMATCH',
                            severity='FLAG',
                            message=f'Pension mismatch: expected ETB {expected:,.2f}, got ETB {emp.pension_employee:,.2f}',
                            hint='Pension should be 7% of basic salary.',
                        )
                    )

        # Check salary typos (>500k ETB)
        for emp in employees:
            total = emp.basic_salary + emp.allowances
            if total > 500000:
                exceptions.append(
                    ExceptionItem(
                        employee_id=emp.employee_id,
                        employee_name=emp.employee_name,
                        rule_code='SALARY_HIGH',
                        severity='FLAG',
                        message=f'Unusually high salary: ETB {total:,.2f}',
                        hint='Verify this is correct.',
                    )
                )

        # Check salary changes from previous period
        if previous_payslips:
            for emp in employees:
                if emp.employee_id in previous_payslips:
                    prev = previous_payslips[emp.employee_id]
                    prev_total = _D(prev.get('basic', 0)) + _D(prev.get('allowances', 0))
                    curr_total = emp.basic_salary + emp.allowances
                    if prev_total > 0:
                        change_pct = abs(float(curr_total - prev_total) / float(prev_total) * 100)
                        if change_pct > 30:
                            direction = 'increased' if curr_total > prev_total else 'decreased'
                            exceptions.append(
                                ExceptionItem(
                                    employee_id=emp.employee_id,
                                    employee_name=emp.employee_name,
                                    rule_code='SALARY_CHANGE_30PCT',
                                    severity='FLAG',
                                    message=f'Salary {direction} by {change_pct:.0f}% (ETB {prev_total:,.0f} → {curr_total:,.0f})',
                                    hint='Verify this salary change is correct.',
                                )
                            )

        # Check missing TIN
        for emp in employees:
            if not emp.tin.strip():
                exceptions.append(
                    ExceptionItem(
                        employee_id=emp.employee_id,
                        employee_name=emp.employee_name,
                        rule_code='MISSING_TIN',
                        severity='WARN',
                        message='No TIN number — required for ERCA filing',
                        hint='Ask the employee for their TIN.',
                    )
                )

        # Cash compliance (>50k requires electronic payment)
        for emp in employees:
            if emp.net > 50000 and not emp.bank_account.strip():
                exceptions.append(
                    ExceptionItem(
                        employee_id=emp.employee_id,
                        employee_name=emp.employee_name,
                        rule_code='CASH_COMPLIANCE',
                        severity='FLAG',
                        message=f'Net pay ETB {emp.net:,.2f} exceeds ETB 50,000 cash limit',
                        hint='Electronic payment required by law (Proclamation 1395/2025, Art. 81)',
                    )
                )

        return exceptions

    # ------------------------------------------------------------------
    # Change detection
    # ------------------------------------------------------------------

    def _compute_changes(self, run: PayrollRunResult, previous_payslips: dict):
        """Compute changes vs previous period."""
        current_map = {e.employee_id: e for e in run.employees}
        prev_ids = set(previous_payslips.keys())
        curr_ids = set(current_map.keys())

        # New hires
        for eid in curr_ids - prev_ids:
            emp = current_map[eid]
            run.new_hires.append(
                {
                    'employee_id': eid,
                    'name': emp.employee_name,
                    'gross': emp.gross,
                }
            )

        # Departures
        for eid in prev_ids - curr_ids:
            prev = previous_payslips[eid]
            run.departures.append(
                {
                    'employee_id': eid,
                    'name': prev.get('name', ''),
                    'last_gross': _D(prev.get('basic', 0)) + _D(prev.get('allowances', 0)),
                }
            )

        # Salary changes
        for eid in curr_ids & prev_ids:
            emp = current_map[eid]
            prev = previous_payslips[eid]
            prev_gross = _D(prev.get('basic', 0)) + _D(prev.get('allowances', 0))
            if emp.gross != prev_gross:
                delta = emp.gross - prev_gross
                pct = float(delta / prev_gross * 100) if prev_gross > 0 else 0
                run.salary_changes.append(
                    {
                        'employee_id': eid,
                        'name': emp.employee_name,
                        'old_gross': prev_gross,
                        'new_gross': emp.gross,
                        'delta': delta,
                        'delta_pct': round(pct, 1),
                    }
                )

        # Previous totals
        run.previous_employee_count = len(previous_payslips)
        run.previous_total_gross = sum(
            (_D(p.get('basic', 0)) + _D(p.get('allowances', 0))) for p in previous_payslips.values()
        )
        run.previous_total_net = sum(_D(p.get('net', 0)) for p in previous_payslips.values())

        if run.previous_total_gross > 0:
            run.gross_delta = run.total_gross - run.previous_total_gross
            run.gross_delta_pct = (run.gross_delta / run.previous_total_gross * 100).quantize(Q)

    # ------------------------------------------------------------------
    # Bank file generation
    # ------------------------------------------------------------------

    def _generate_bank_file(self, run: PayrollRunResult):
        """Generate bank-ready CSV from approved payroll."""
        from payroll_engine.bank_file import generate_csv

        employees_data = []
        for emp in run.employees:
            if emp.status != 'exception':
                employees_data.append(
                    {
                        'id': emp.employee_id,
                        'name': emp.employee_name,
                        'bank': emp.bank_account,
                        'net': float(emp.net),
                    }
                )

        if employees_data:
            csv_bytes = generate_csv(
                employees_data,
                bank='cbe',
                company_name=run.company_name,
                period=run.period,
            )
            run.bank_file_data = csv_bytes
            run.bank_file_name = f'bank_transfer_{run.period or "payroll"}.csv'

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    def export_to_excel(self, run: PayrollRunResult) -> bytes:
        """
        Export complete payroll results to a multi-sheet Excel workbook.

        Sheets:
        1. Summary — totals, period info, approval status
        2. Payroll — per-employee calculation results
        3. Calculation Flow — step-by-step for each employee
        4. Tax Breakdown — bracket-by-bracket detail
        5. Exceptions — all BLOCK/FLAG/WARN items
        6. Changes — vs previous period
        7. Bank File — ready for upload
        8. Approval — signature block
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Alignment,
                Border,
                Font,
                PatternFill,
                Side,
                numbers,
            )
        except ImportError:
            raise ImportError('openpyxl is required for Excel export')

        wb = openpyxl.Workbook()

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
        title_font = Font(bold=True, size=16, color='1A5276')
        subtitle_font = Font(bold=True, size=12, color='333333')
        etb_format = '#,##0.00'
        pct_format = '0.0%'
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )
        block_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
        flag_fill = PatternFill(start_color='FEF9E7', end_color='FEF9E7', fill_type='solid')
        warn_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
        green_fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
        totals_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
        totals_font = Font(bold=True, size=11, color='1A5276')

        def _style_header(ws, row, col_count):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

        def _style_data_cell(cell, is_number=False):
            cell.border = thin_border
            if is_number:
                cell.number_format = etb_format
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # ---- Sheet 1: Summary ----
        ws = wb.active
        ws.title = 'Summary'
        ws.merge_cells('A1:F1')
        ws['A1'] = run.company_name or 'Payroll Summary'
        ws['A1'].font = title_font

        summary_data = [
            ('Period', run.period or 'N/A'),
            ('Run Date', run.run_date),
            ('Status', run.status.upper()),
            ('Employee Count', run.employee_count),
            ('', ''),
            ('Total Gross', float(run.total_gross)),
            ('Total Tax', float(run.total_tax)),
            ('Total Pension (Employee)', float(run.total_pension_employee)),
            ('Total Pension (Employer)', float(run.total_pension_employer)),
            ('Total Overtime', float(run.total_overtime)),
            ('Total Deductions', float(run.total_deductions)),
            ('Total Net', float(run.total_net)),
            ('', ''),
            ('BLOCK Exceptions', run.block_count),
            ('FLAG Exceptions', run.flag_count),
            ('WARN Exceptions', run.warn_count),
            ('', ''),
            ('Input Hash', run.input_hash),
            ('Calculation Hash', run.calculation_hash),
        ]

        for i, (label, value) in enumerate(summary_data, 3):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True) if label else Font()
            cell = ws.cell(row=i, column=2, value=value)
            if isinstance(value, float):
                cell.number_format = etb_format

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        # ---- Sheet 2: Payroll ----
        ws2 = wb.create_sheet('Payroll')
        headers = [
            'Employee ID', 'Name', 'Department', 'Position',
            'Basic Salary', 'Allowances', 'Exempt Allow.', 'Taxable Allow.',
            'Overtime', 'Gross', 'Pension (7%)', 'Taxable Income',
            'Tax', 'Deductions', 'Net Pay', 'Eff. Tax Rate', 'Status',
        ]
        for col, h in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=h)
        _style_header(ws2, 1, len(headers))

        for i, emp in enumerate(run.employees, 2):
            values = [
                emp.employee_id, emp.employee_name, emp.department, emp.position,
                float(emp.basic_salary), float(emp.allowances),
                float(emp.exempt_allowances), float(emp.taxable_allowances),
                float(emp.overtime_pay), float(emp.gross),
                float(emp.pension_employee), float(emp.taxable),
                float(emp.tax), float(emp.total_deductions),
                float(emp.net), float(emp.effective_tax_rate) / 100,
                emp.status,
            ]
            for col, val in enumerate(values, 1):
                cell = ws2.cell(row=i, column=col, value=val)
                _style_data_cell(cell, is_number=col >= 5 and col <= 16)
                if col == 16:
                    cell.number_format = pct_format
                if col == 17:
                    if val == 'exception':
                        cell.fill = block_fill
                    elif val == 'approved':
                        cell.fill = green_fill

        # Totals row
        totals_row = len(run.employees) + 2
        ws2.cell(row=totals_row, column=1, value='TOTALS').font = totals_font
        ws2.cell(row=totals_row, column=1).fill = totals_fill
        for col in range(2, len(headers) + 1):
            ws2.cell(row=totals_row, column=col).fill = totals_fill
            ws2.cell(row=totals_row, column=col).font = totals_font

        total_cols = {
            5: float(run.total_gross - run.total_overtime - sum((e.taxable_allowances + e.exempt_allowances) for e in run.employees)),
            6: float(sum((e.taxable_allowances + e.exempt_allowances) for e in run.employees)),
            9: float(run.total_overtime),
            10: float(run.total_gross),
            11: float(run.total_pension_employee),
            13: float(run.total_tax),
            14: float(run.total_deductions),
            15: float(run.total_net),
        }
        for col, val in total_cols.items():
            cell = ws2.cell(row=totals_row, column=col, value=val)
            cell.number_format = etb_format
            cell.fill = totals_fill
            cell.font = totals_font

        # Auto-width
        from openpyxl.utils import get_column_letter
        for col in range(1, len(headers) + 1):
            ws2.column_dimensions[get_column_letter(col)].width = 15

        # ---- Sheet 3: Calculation Flow ----
        ws3 = wb.create_sheet('Calculation Flow')
        flow_headers = ['Employee ID', 'Name', 'Step', 'Label', 'Formula', 'Inputs', 'Result', 'Note', 'Legal Reference']
        for col, h in enumerate(flow_headers, 1):
            ws3.cell(row=1, column=col, value=h)
        _style_header(ws3, 1, len(flow_headers))

        row = 2
        for emp in run.employees:
            for step in emp.steps:
                ws3.cell(row=row, column=1, value=emp.employee_id)
                ws3.cell(row=row, column=2, value=emp.employee_name)
                ws3.cell(row=row, column=3, value=step.step_number)
                ws3.cell(row=row, column=4, value=step.label)
                ws3.cell(row=row, column=5, value=step.formula)
                ws3.cell(row=row, column=6, value=json.dumps({k: str(v) for k, v in step.inputs.items()}))
                cell = ws3.cell(row=row, column=7, value=float(step.result))
                cell.number_format = etb_format
                ws3.cell(row=row, column=8, value=step.note)
                ws3.cell(row=row, column=9, value=step.legal_reference)
                for col in range(1, len(flow_headers) + 1):
                    ws3.cell(row=row, column=col).border = thin_border
                row += 1

        # ---- Sheet 4: Tax Breakdown ----
        ws4 = wb.create_sheet('Tax Breakdown')
        tax_headers = ['Employee ID', 'Name', 'Taxable Income', 'Bracket', 'Rate', 'Amount in Bracket', 'Tax']
        for col, h in enumerate(tax_headers, 1):
            ws4.cell(row=1, column=col, value=h)
        _style_header(ws4, 1, len(tax_headers))

        row = 2
        for emp in run.employees:
            if emp.tax_breakdown and emp.tax_breakdown.get('brackets'):
                for bracket in emp.tax_breakdown['brackets']:
                    ws4.cell(row=row, column=1, value=emp.employee_id)
                    ws4.cell(row=row, column=2, value=emp.employee_name)
                    ws4.cell(row=row, column=3, value=float(emp.taxable))
                    upper = bracket.get('upper')
                    bracket_label = f"ETB {bracket['lower']:,.0f} – {upper:,.0f}" if upper else f"ETB {bracket['lower']:,.0f}+"
                    ws4.cell(row=row, column=4, value=bracket_label)
                    ws4.cell(row=row, column=5, value=f"{bracket['rate_pct']}%")
                    ws4.cell(row=row, column=6, value=float(bracket['taxable_amount']))
                    ws4.cell(row=row, column=7, value=float(bracket['bracket_tax']))
                    for col in range(1, len(tax_headers) + 1):
                        cell = ws4.cell(row=row, column=col)
                        cell.border = thin_border
                        if col in (3, 6, 7):
                            cell.number_format = etb_format
                    row += 1

        # ---- Sheet 5: Exceptions ----
        ws5 = wb.create_sheet('Exceptions')
        exc_headers = ['Severity', 'Employee ID', 'Name', 'Rule Code', 'Message', 'Hint', 'Overridden', 'Override Reason']
        for col, h in enumerate(exc_headers, 1):
            ws5.cell(row=1, column=col, value=h)
        _style_header(ws5, 1, len(exc_headers))

        for i, exc in enumerate(run.exceptions, 2):
            values = [
                exc.severity, exc.employee_id, exc.employee_name,
                exc.rule_code, exc.message, exc.hint,
                'Yes' if exc.overridden else 'No', exc.override_reason,
            ]
            for col, val in enumerate(values, 1):
                cell = ws5.cell(row=i, column=col, value=val)
                cell.border = thin_border
                if exc.severity == 'BLOCK':
                    cell.fill = block_fill
                elif exc.severity == 'FLAG':
                    cell.fill = flag_fill
                elif exc.severity == 'WARN':
                    cell.fill = warn_fill

        # ---- Sheet 6: Changes ----
        ws6 = wb.create_sheet('Changes')
        ws6.merge_cells('A1:E1')
        ws6['A1'] = f'Changes vs Previous Period'
        ws6['A1'].font = subtitle_font

        if run.previous_employee_count > 0:
            ws6.cell(row=3, column=1, value='Previous Period').font = Font(bold=True)
            ws6.cell(row=3, column=2, value=f'{run.previous_employee_count} employees')
            ws6.cell(row=4, column=1, value='Previous Gross').font = Font(bold=True)
            ws6.cell(row=4, column=2, value=float(run.previous_total_gross)).number_format = etb_format
            ws6.cell(row=5, column=1, value='Gross Change').font = Font(bold=True)
            ws6.cell(row=5, column=2, value=float(run.gross_delta)).number_format = etb_format
            ws6.cell(row=5, column=3, value=f'{run.gross_delta_pct}%')

        # New hires
        row = 7
        if run.new_hires:
            ws6.cell(row=row, column=1, value='NEW HIRES').font = Font(bold=True, color='27AE60')
            row += 1
            for h in run.new_hires:
                ws6.cell(row=row, column=1, value=h['employee_id'])
                ws6.cell(row=row, column=2, value=h['name'])
                ws6.cell(row=row, column=3, value=float(h['gross'])).number_format = etb_format
                row += 1

        # Departures
        if run.departures:
            row += 1
            ws6.cell(row=row, column=1, value='DEPARTURES').font = Font(bold=True, color='E74C3C')
            row += 1
            for d in run.departures:
                ws6.cell(row=row, column=1, value=d['employee_id'])
                ws6.cell(row=row, column=2, value=d['name'])
                ws6.cell(row=row, column=3, value=float(d['last_gross'])).number_format = etb_format
                row += 1

        # Salary changes
        if run.salary_changes:
            row += 1
            ws6.cell(row=row, column=1, value='SALARY CHANGES').font = Font(bold=True, color='F39C12')
            row += 1
            change_headers = ['ID', 'Name', 'Old Gross', 'New Gross', 'Change', '%']
            for col, h in enumerate(change_headers, 1):
                ws6.cell(row=row, column=col, value=h)
            _style_header(ws6, row, len(change_headers))
            row += 1
            for c in run.salary_changes:
                ws6.cell(row=row, column=1, value=c['employee_id'])
                ws6.cell(row=row, column=2, value=c['name'])
                ws6.cell(row=row, column=3, value=float(c['old_gross'])).number_format = etb_format
                ws6.cell(row=row, column=4, value=float(c['new_gross'])).number_format = etb_format
                ws6.cell(row=row, column=5, value=float(c['delta'])).number_format = etb_format
                ws6.cell(row=row, column=6, value=f"{c['delta_pct']}%")
                row += 1

        # ---- Sheet 7: Bank File ----
        ws7 = wb.create_sheet('Bank File')
        if run.bank_file_data:
            ws7.cell(row=1, column=1, value='Bank Transfer File').font = subtitle_font
            ws7.cell(row=2, column=1, value=f'Period: {run.period}')
            ws7.cell(row=3, column=1, value=f'Employees: {run.employee_count}')
            ws7.cell(row=4, column=1, value=f'Total: ETB {run.total_net:,.2f}')

            bank_headers = ['Account Number', 'Amount', 'Narrative', 'Currency']
            for col, h in enumerate(bank_headers, 1):
                ws7.cell(row=6, column=col, value=h)
            _style_header(ws7, 6, len(bank_headers))

            # Parse the CSV bank file
            import csv
            import io

            reader = csv.reader(io.StringIO(run.bank_file_data.decode('utf-8')))
            next(reader)  # skip header
            for i, row_data in enumerate(reader, 7):
                for col, val in enumerate(row_data, 1):
                    cell = ws7.cell(row=i, column=col, value=val)
                    cell.border = thin_border
                    if col == 1:
                        cell.number_format = '@'  # TEXT format
                    elif col == 2:
                        try:
                            cell.value = float(val)
                            cell.number_format = etb_format
                        except ValueError:
                            pass
        else:
            ws7.cell(row=1, column=1, value='No bank file generated').font = Font(italic=True, color='E74C3C')
            ws7.cell(row=2, column=1, value='Resolve BLOCK exceptions first')

        # ---- Sheet 8: Approval ----
        ws8 = wb.create_sheet('Approval')
        ws8.merge_cells('A1:D1')
        ws8['A1'] = 'Payroll Approval'
        ws8['A1'].font = title_font

        approval_data = [
            ('Period', run.period),
            ('Employee Count', run.employee_count),
            ('Total Net Pay', f'ETB {run.total_net:,.2f}'),
            ('Exceptions', f'{run.block_count} BLOCK, {run.flag_count} FLAG, {run.warn_count} WARN'),
            ('Status', run.status.upper()),
            ('', ''),
            ('Approved By', run.approved_by or '_________________________'),
            ('Date', run.approved_at or '____/____/________'),
            ('Signature', ''),
            ('', ''),
            ('Notes', run.approval_notes or ''),
            ('', ''),
            ('Calculation Hash', run.calculation_hash),
            ('Input Hash', run.input_hash),
        ]

        for i, (label, value) in enumerate(approval_data, 3):
            ws8.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws8.cell(row=i, column=2, value=str(value))

        ws8.column_dimensions['A'].width = 20
        ws8.column_dimensions['B'].width = 40

        # Save
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # ------------------------------------------------------------------
    # Approval workflow
    # ------------------------------------------------------------------

    def approve(self, run: PayrollRunResult, approved_by: str, notes: str = '') -> PayrollRunResult:
        """Approve a payroll run. Blocks if there are unresolved BLOCK exceptions."""
        if run.block_count > 0:
            raise ValueError(
                f'Cannot approve: {run.block_count} BLOCK exception(s) unresolved. '
                f'Resolve them or override with a reason.'
            )

        run.status = 'approved'
        run.approved_by = approved_by
        run.approved_at = datetime.now(UTC).isoformat()
        run.approval_notes = notes

        # Mark all employees as approved
        for emp in run.employees:
            if emp.status != 'exception':
                emp.status = 'approved'

        return run

    def override_exception(
        self,
        run: PayrollRunResult,
        employee_id: str,
        rule_code: str,
        reason: str,
        overridden_by: str,
    ) -> PayrollRunResult:
        """Override a FLAG exception with a reason."""
        for exc in run.exceptions:
            if exc.employee_id == employee_id and exc.rule_code == rule_code:
                if exc.severity == 'BLOCK':
                    raise ValueError('BLOCK exceptions cannot be overridden. Fix the issue first.')
                exc.overridden = True
                exc.override_reason = reason
                exc.overridden_by = overridden_by

                # Recalculate exception counts
                run.block_count = sum(1 for e in run.exceptions if e.severity == 'BLOCK' and not e.overridden)
                run.flag_count = sum(1 for e in run.exceptions if e.severity == 'FLAG' and not e.overridden)

                # Update employee status if all exceptions resolved
                for emp in run.employees:
                    if emp.employee_id == employee_id:
                        unresolved = [e for e in emp.exceptions if not e.overridden and e.severity == 'BLOCK']
                        if not unresolved:
                            emp.status = 'calculated'
                            emp.exception_reason = ''

                break

        return run

    def lock(self, run: PayrollRunResult, locked_by: str) -> PayrollRunResult:
        """Lock a completed payroll run. No further changes allowed."""
        if run.status not in ('approved', 'completed'):
            raise ValueError(f'Cannot lock: status is {run.status}. Must be approved first.')
        run.status = 'locked'
        return run
