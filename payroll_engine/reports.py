"""
Report Generation Module

Generates downloadable reports for:
- ERCA tax filing (monthly)
- Pension contribution reporting (monthly)
- Year-end reconciliation

All reports are generated from Payslip data and downloadable as Excel.
"""

import io
from datetime import date
from typing import List, Dict, Any


def generate_erca_report(payslips: list, company_name: str,
                         period: str = None, company=None) -> bytes:
    """
    Generate ERCA-formatted tax filing report.

    Uses company's report template if available, otherwise uses default columns.

    Args:
        payslips: List of Payslip objects with employee relationship loaded
        company_name: Company name for the report header
        period: Period string (e.g., "July 2025"), defaults to current month
        company: Company instance (for template-aware column configuration)

    Returns:
        Excel file as bytes
    """
    from payroll_engine.report_templates import get_enabled_columns, get_column_value

    # Get columns from template or use defaults (ERCA portal format)
    if company:
        columns = get_enabled_columns(company, 'erca')
    else:
        # Fallback: use ERCA portal columns
        columns = [
            {'key': 'employee_name', 'label': 'Employee Full Name', 'data_path': 'employee.name'},
            {'key': 'start_date', 'label': 'Start Date', 'data_path': 'employee.start_date'},
            {'key': 'end_date', 'label': 'End Date', 'data_path': '_end_date'},
            {'key': 'basic_salary', 'label': 'Basic Salary', 'data_path': 'employee.basic_salary'},
            {'key': 'transport_allowance', 'label': 'Transport Allowance', 'data_path': '_transport_allowance'},
            {'key': 'taxable_transport', 'label': 'Taxable Transport Allowance', 'data_path': '_taxable_transport'},
            {'key': 'overtime_pay', 'label': 'Over Time', 'data_path': 'overtime_pay'},
            {'key': 'other_taxable', 'label': 'Other Taxable Benefit', 'data_path': '_other_taxable'},
            {'key': 'total_taxable', 'label': 'Total Taxable', 'data_path': 'taxable'},
            {'key': 'tax_withheld', 'label': 'Tax withheld', 'data_path': 'tax'},
        ]

    # Identify which columns are numeric (for formatting and totals)
    NUMERIC_KEYS = {
        'basic_salary', 'allowances', 'gross_salary', 'overtime_pay',
        'pension_employee', 'pension_employer', 'taxable_income',
        'tax_withheld', 'net_pay', 'transport_allowance',
        'taxable_transport', 'other_taxable', 'total_taxable',
    }

    # Map data_path to actual payslip attribute names
    PATH_MAP = {
        '_row_number': '_row_number',
        '_end_date': '_end_date',
        '_transport_allowance': '_transport_allowance',
        '_taxable_transport': '_taxable_transport',
        '_other_taxable': '_other_taxable',
        'employee.employee_id': 'employee.employee_id',
        'employee.name': 'employee.name',
        'employee.tin': 'employee.tin',
        'employee.start_date': 'employee.start_date',
        'employee.department': 'employee.department',
        'employee.position': 'employee.position',
        'employee.basic_salary': 'employee.basic_salary',
        'employee.allowances': 'employee.allowances',
        'gross': 'gross_salary',
        'overtime_pay': 'overtime_pay',
        'pension_employee': 'employee_pension',
        'pension_employer': 'employer_pension',
        'taxable': 'taxable',
        'tax': 'tax',
        'net': 'net_pay',
        '_company_tin': '_company_tin',
        '_company_name': '_company_name',
        'employee.bank_account': 'employee.bank_account',
        'employee.bank_or_telebirr': 'employee.bank_or_telebirr',
    }

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        return _generate_erca_csv(payslips, company_name, period)

    if period is None:
        period = date.today().strftime('%B %Y')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ERCA Tax Filing"

    # Styles
    title_font = Font(bold=True, size=16, color='1A5276')
    subtitle_font = Font(bold=True, size=12, color='333333')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_align = Alignment(horizontal='right', vertical='center')
    name_align = Alignment(horizontal='left', vertical='center')
    totals_font = Font(bold=True, size=11, color='1A5276')
    totals_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    etb_format = '#,##0.00'

    # --- Column headers (row 5) ---
    num_cols = len(columns)
    col_letter_last = openpyxl.utils.get_column_letter(num_cols)

    # Merge title cells to span all columns
    ws.merge_cells(f'A1:{col_letter_last}1')
    ws['A1'] = company_name
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'A2:{col_letter_last}2')
    ws['A2'] = f'ERCA Monthly Tax Filing Report — {period}'
    ws['A2'].font = subtitle_font

    ws.merge_cells(f'A3:{col_letter_last}3')
    ws['A3'] = f'Generated: {date.today().strftime("%d %B %Y")}'
    ws['A3'].font = Font(italic=True, color='666666')

    # Auto-calculate column widths based on label length
    for col_idx, col_cfg in enumerate(columns, 1):
        width = max(len(col_cfg['label']) + 2, 10)
        width = min(width, 30)  # cap at 30
        cell = ws.cell(row=5, column=col_idx, value=col_cfg['label'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # --- Data rows ---
    totals = {}  # key -> running total

    for i, p in enumerate(payslips, 1):
        emp = p.employee
        taxable = p.gross_salary - p.employee_pension
        row = 5 + i

        for col_idx, col_cfg in enumerate(columns, 1):
            key = col_cfg['key']
            data_path = col_cfg.get('data_path', key)
            static_value = col_cfg.get('static_value')

            # Extract value — use get_column_value for flexibility
            val = get_column_value(p, data_path, company, static_value)

            # Handle row number specially
            if data_path == '_row_number':
                val = i

            if val is None:
                val = ''

            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border

            # Format numeric columns
            if key in NUMERIC_KEYS and isinstance(val, (int, float)):
                cell.number_format = etb_format
                cell.alignment = data_align
                totals[key] = totals.get(key, 0) + val
            elif key == 'row_number':
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = name_align

    # --- Totals row ---
    totals_row = 5 + len(payslips) + 1
    # Find the first text column for the TOTALS label
    label_col = 1
    for col_idx, col_cfg in enumerate(columns, 1):
        if col_cfg['key'] in ('employee_name', 'employee_id'):
            label_col = col_idx
            break

    ws.cell(row=totals_row, column=label_col, value='TOTALS').font = totals_font
    ws.cell(row=totals_row, column=label_col).fill = totals_fill
    ws.cell(row=totals_row, column=label_col).border = thin_border

    for col_idx, col_cfg in enumerate(columns, 1):
        key = col_cfg['key']
        if key in totals:
            cell = ws.cell(row=totals_row, column=col_idx, value=totals[key])
            cell.font = totals_font
            cell.fill = totals_fill
            cell.number_format = etb_format
            cell.alignment = data_align
            cell.border = thin_border

    # --- Print setup ---
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = 'landscape'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_pension_report(payslips: list, company_name: str,
                            period: str = None) -> bytes:
    """
    Generate pension contribution report.

    Args:
        payslips: List of Payslip objects with employee relationship loaded
        company_name: Company name for the report header
        period: Period string (e.g., "July 2025")

    Returns:
        Excel file as bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        return _generate_pension_csv(payslips, company_name, period)

    if period is None:
        period = date.today().strftime('%B %Y')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pension Report"

    # Styles
    title_font = Font(bold=True, size=16, color='1A5276')
    subtitle_font = Font(bold=True, size=12, color='333333')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_align = Alignment(horizontal='right', vertical='center')
    name_align = Alignment(horizontal='left', vertical='center')
    totals_font = Font(bold=True, size=11, color='1A5276')
    totals_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    etb_format = '#,##0.00'

    # --- Title block ---
    ws.merge_cells('A1:G1')
    ws['A1'] = company_name
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('A2:G2')
    ws['A2'] = f'Pension Contribution Report — {period}'
    ws['A2'].font = subtitle_font

    ws.merge_cells('A3:G3')
    ws['A3'] = f'Generated: {date.today().strftime("%d %B %Y")}'
    ws['A3'].font = Font(italic=True, color='666666')

    # --- Column headers (row 5) ---
    headers = [
        'No.', 'Employee ID', 'Employee Name', 'Basic Salary',
        'Employee 7%', 'Employer 11%', 'Total'
    ]
    col_widths = [6, 14, 22, 16, 14, 14, 16]
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # --- Data rows ---
    total_employee = 0
    total_employer = 0

    for i, p in enumerate(payslips, 1):
        emp = p.employee
        total = p.employee_pension + p.employer_pension
        row = 5 + i

        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=emp.employee_id).alignment = name_align
        ws.cell(row=row, column=3, value=emp.name).alignment = name_align

        for col_idx, val in [(4, emp.basic_salary), (5, p.employee_pension),
                              (6, p.employer_pension), (7, total)]:
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.number_format = etb_format
            cell.alignment = data_align

        for col_idx in range(1, 8):
            ws.cell(row=row, column=col_idx).border = thin_border

        total_employee += p.employee_pension
        total_employer += p.employer_pension

    # --- Totals row ---
    totals_row = 5 + len(payslips) + 1
    ws.cell(row=totals_row, column=3, value='TOTALS').font = totals_font
    ws.cell(row=totals_row, column=3).fill = totals_fill
    ws.cell(row=totals_row, column=3).border = thin_border
    for col_idx, val in [(5, total_employee), (6, total_employer),
                          (7, total_employee + total_employer)]:
        cell = ws.cell(row=totals_row, column=col_idx, value=val)
        cell.font = totals_font
        cell.fill = totals_fill
        cell.number_format = etb_format
        cell.alignment = data_align
        cell.border = thin_border

    # --- Print setup ---
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = 'landscape'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_yearly_summary(payslips: list, company_name: str, year: int) -> bytes:
    """Generate annual tax/pension summary per employee for a given year.

    Args:
        payslips: All Payslip objects for the year (with employee loaded).
        company_name: Company name for header.
        year: Calendar year (e.g. 2026).

    Returns:
        Excel file as bytes.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return _generate_yearly_csv(payslips, company_name, year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Year-End {year}'

    title_font = Font(bold=True, size=16, color='1A5276')
    subtitle_font = Font(bold=True, size=12, color='333333')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    totals_font = Font(bold=True, size=11)
    totals_fill = PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid')
    etb_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    name_align = Alignment(vertical='center')
    data_align = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('A1:H1')
    ws['A1'] = f'{company_name}'
    ws['A1'].font = title_font

    ws.merge_cells('A2:H2')
    ws['A2'] = f'Year-End Summary — {year}'
    ws['A2'].font = subtitle_font

    ws.merge_cells('A3:H3')
    ws['A3'] = f'Generated: {date.today().strftime("%d %B %Y")}'
    ws['A3'].font = Font(italic=True, color='666666')

    headers = [
        'No.', 'Employee ID', 'Employee Name', 'Total Gross',
        'Total Pension', 'Total Tax', 'Total Net', 'Periods Paid',
    ]
    col_widths = [6, 14, 22, 16, 14, 14, 16, 12]
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    from collections import defaultdict
    emp_agg = defaultdict(lambda: {'gross': 0, 'pension': 0, 'tax': 0, 'net': 0, 'periods': set()})
    for p in payslips:
        eid = p.employee_id
        emp_agg[eid]['gross'] += p.gross_salary
        emp_agg[eid]['pension'] += p.employee_pension
        emp_agg[eid]['tax'] += p.tax
        emp_agg[eid]['net'] += p.net_pay
        emp_agg[eid]['periods'].add(p.payroll_run_id)
        if not emp_agg[eid].get('name'):
            emp_agg[eid]['name'] = p.employee.name if p.employee else f'Employee {eid}'
        if not emp_agg[eid].get('emp_id'):
            emp_agg[eid]['emp_id'] = p.employee.employee_id if p.employee else str(eid)

    grand_gross = grand_pension = grand_tax = grand_net = 0
    for i, (eid, agg) in enumerate(sorted(emp_agg.items()), 1):
        row = 5 + i
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=agg.get('emp_id', eid)).alignment = name_align
        ws.cell(row=row, column=3, value=agg.get('name', '')).alignment = name_align
        for col_idx, val in [(4, agg['gross']), (5, agg['pension']),
                              (6, agg['tax']), (7, agg['net'])]:
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.number_format = etb_format
            cell.alignment = data_align
        ws.cell(row=row, column=8, value=len(agg['periods'])).alignment = Alignment(horizontal='center')
        for col_idx in range(1, 9):
            ws.cell(row=row, column=col_idx).border = thin_border
        grand_gross += agg['gross']
        grand_pension += agg['pension']
        grand_tax += agg['tax']
        grand_net += agg['net']

    totals_row = 5 + len(emp_agg) + 1
    ws.cell(row=totals_row, column=3, value='TOTALS').font = totals_font
    ws.cell(row=totals_row, column=3).fill = totals_fill
    ws.cell(row=totals_row, column=3).border = thin_border
    for col_idx, val in [(4, grand_gross), (5, grand_pension), (6, grand_tax), (7, grand_net)]:
        cell = ws.cell(row=totals_row, column=col_idx, value=val)
        cell.font = totals_font
        cell.fill = totals_fill
        cell.number_format = etb_format
        cell.alignment = data_align
        cell.border = thin_border

    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = 'landscape'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _generate_yearly_csv(payslips, company_name, year):
    """Fallback CSV for year-end summary."""
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Year-End Summary', company_name, str(year)])
    writer.writerow(['No.', 'Employee ID', 'Name', 'Total Gross', 'Total Pension', 'Total Tax', 'Total Net', 'Periods Paid'])

    from collections import defaultdict
    emp_agg = defaultdict(lambda: {'gross': 0, 'pension': 0, 'tax': 0, 'net': 0, 'periods': set()})
    for p in payslips:
        eid = p.employee_id
        emp_agg[eid]['gross'] += p.gross_salary
        emp_agg[eid]['pension'] += p.employee_pension
        emp_agg[eid]['tax'] += p.tax
        emp_agg[eid]['net'] += p.net_pay
        emp_agg[eid]['periods'].add(p.payroll_run_id)
        if not emp_agg[eid].get('name'):
            emp_agg[eid]['name'] = p.employee.name if p.employee else f'Employee {eid}'
        if not emp_agg[eid].get('emp_id'):
            emp_agg[eid]['emp_id'] = p.employee.employee_id if p.employee else str(eid)

    from payroll_engine.security import prevent_csv_injection
    for i, (eid, agg) in enumerate(sorted(emp_agg.items()), 1):
        writer.writerow([
            i,
            prevent_csv_injection(agg.get('emp_id', eid)),
            prevent_csv_injection(agg.get('name', '')),
            agg['gross'], agg['pension'], agg['tax'], agg['net'], len(agg['periods']),
        ])
    return output.getvalue().encode('utf-8')


def _generate_erca_csv(payslips, company_name, period):
    """Fallback CSV generation if openpyxl not installed.
    Uses ERCA portal column format.
    """
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ERCA Tax Filing Report', company_name, period or ''])
    # ERCA portal headers
    writer.writerow([
        'Employee Full Name', 'Start Date', 'End Date',
        'Basic Salary', 'Transport Allowance', 'Taxable Transport Allowance',
        'Over Time', 'Other Taxable Benefit', 'Total Taxable', 'Tax withheld',
    ])
    for i, p in enumerate(payslips, 1):
        emp = p.employee
        basic = float(emp.basic_salary or 0)
        gross = float(p.gross_salary or 0)
        transport = max(0, gross - basic)
        start_date = emp.start_date.strftime('%d/%m/%Y') if emp.start_date else ''
        from payroll_engine.security import prevent_csv_injection
        writer.writerow([
            prevent_csv_injection(emp.name),
            start_date,
            '',  # End Date
            basic,
            transport,  # Transport Allowance
            transport,  # Taxable Transport Allowance
            float(p.overtime_pay or 0),  # Over Time
            0,  # Other Taxable Benefit
            float(p.taxable or 0),  # Total Taxable
            float(p.tax or 0),  # Tax withheld
        ])
    return output.getvalue().encode('utf-8')


def _generate_pension_csv(payslips, company_name, period):
    """Fallback CSV generation if openpyxl not installed."""
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Pension Contribution Report', company_name, period or ''])
    writer.writerow(['No.', 'Employee ID', 'Name', 'Basic Salary', 'Employee 7%', 'Employer 11%', 'Total'])
    for i, p in enumerate(payslips, 1):
        emp = p.employee
        total = p.employee_pension + p.employer_pension
        from payroll_engine.security import prevent_csv_injection
        writer.writerow([
            i,
            prevent_csv_injection(emp.employee_id),
            prevent_csv_injection(emp.name),
            emp.basic_salary, p.employee_pension, p.employer_pension, total,
        ])
    return output.getvalue().encode('utf-8')
