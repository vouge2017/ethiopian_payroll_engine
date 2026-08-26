"""
Tests for the Excel-Compatible Payroll Engine.

Tests cover:
1. Deterministic calculation (same inputs → same outputs)
2. Calculation flow completeness
3. Tax bracket breakdown
4. Exception detection
5. Change detection
6. Approval workflow
7. Bank file generation
8. Excel export
9. Edge cases
"""

import io
import json
import os
import tempfile
from decimal import Decimal

import pytest

from payroll_engine.excel_payroll import (
    CalculationStep,
    EmployeePayrollResult,
    ExcelPayrollEngine,
    ExceptionItem,
    PayrollRunResult,
    _D,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def engine():
    """Create an engine instance without database."""
    return ExcelPayrollEngine(company_id=1, for_date='2026-08-01')


@pytest.fixture
def sample_employees():
    """Sample employee data for testing."""
    return [
        {
            'employee_id': 'EMP001',
            'name': 'Abebe Kebede',
            'basic_salary': 15000,
            'allowances': 3000,
            'department': 'Engineering',
            'position': 'Senior Developer',
            'bank_account': 'bank:cbe:1000123456789',
            'tin': '0012345678',
        },
        {
            'employee_id': 'EMP002',
            'name': 'Tigist Haile',
            'basic_salary': 8000,
            'allowances': 1500,
            'department': 'Finance',
            'position': 'Accountant',
            'bank_account': 'bank:cbe:1000987654321',
            'tin': '0098765432',
        },
        {
            'employee_id': 'EMP003',
            'name': 'Dawit Tesfaye',
            'basic_salary': 25000,
            'allowances': 5000,
            'department': 'Management',
            'position': 'Director',
            'bank_account': 'telebirr:0911234567',
            'tin': '0055566677',
        },
    ]


# ---------------------------------------------------------------------------
# Test: Deterministic calculation
# ---------------------------------------------------------------------------


class TestDeterminism:
    """Same inputs must always produce the same outputs."""

    def test_same_inputs_same_hash(self, engine, sample_employees):
        """Running the same data twice must produce identical hashes."""
        result1 = engine.run_from_data(sample_employees, period='2018-10')
        result2 = engine.run_from_data(sample_employees, period='2018-10')

        assert result1.calculation_hash == result2.calculation_hash
        assert result1.input_hash == result2.input_hash

    def test_same_inputs_same_totals(self, engine, sample_employees):
        """Totals must be identical across runs."""
        result1 = engine.run_from_data(sample_employees)
        result2 = engine.run_from_data(sample_employees)

        assert result1.total_gross == result2.total_gross
        assert result1.total_tax == result2.total_tax
        assert result1.total_net == result2.total_net
        assert result1.total_pension_employee == result2.total_pension_employee

    def test_same_inputs_same_per_employee(self, engine, sample_employees):
        """Per-employee results must be identical."""
        result1 = engine.run_from_data(sample_employees)
        result2 = engine.run_from_data(sample_employees)

        for e1, e2 in zip(result1.employees, result2.employees):
            assert e1.gross == e2.gross
            assert e1.tax == e2.tax
            assert e1.net == e2.net
            assert e1.pension_employee == e2.pension_employee
            assert e1.taxable == e2.taxable

    def test_decimal_precision(self, engine):
        """Must use Decimal for all calculations (no float drift)."""
        result = engine.calculate_employee(
            employee_id='TEST001',
            name='Precision Test',
            basic_salary=Decimal('12345.67'),
            allowances=Decimal('2345.67'),
        )

        # All values should be Decimal
        assert isinstance(result.gross, Decimal)
        assert isinstance(result.tax, Decimal)
        assert isinstance(result.net, Decimal)
        assert isinstance(result.pension_employee, Decimal)

        # Gross = basic + allowances
        assert result.gross == Decimal('14691.34')

    def test_determinism_with_overtime(self, engine):
        """Overtime calculations must also be deterministic."""
        result1 = engine.calculate_employee(
            employee_id='TEST002',
            name='OT Test',
            basic_salary=Decimal('10000'),
            overtime_hours=Decimal('10'),
            overtime_type='day',
        )
        result2 = engine.calculate_employee(
            employee_id='TEST002',
            name='OT Test',
            basic_salary=Decimal('10000'),
            overtime_hours=Decimal('10'),
            overtime_type='day',
        )

        assert result1.overtime_pay == result2.overtime_pay
        assert result1.net == result2.net


# ---------------------------------------------------------------------------
# Test: Calculation correctness
# ---------------------------------------------------------------------------


class TestCalculation:
    """Verify payroll calculation matches Ethiopian tax law."""

    def test_basic_salary_only(self, engine):
        """Basic salary with no allowances."""
        result = engine.calculate_employee(
            employee_id='E001',
            name='Test',
            basic_salary=Decimal('10000'),
        )

        # Pension: 7% of 10000 = 700
        assert result.pension_employee == Decimal('700.00')
        # Employer pension: 11% of 10000 = 1100
        assert result.pension_employer == Decimal('1100.00')
        # Taxable: 10000 - 700 = 9300
        assert result.taxable == Decimal('9300.00')
        # Tax on 9300:
        #   0-2000: 0
        #   2001-4000: 2000 × 15% = 300
        #   4001-7000: 3000 × 20% = 600
        #   7001-9300: 2300 × 25% = 575
        #   Total = 1475
        assert result.tax == Decimal('1475.00')
        # Net: 10000 - 700 - 1475 = 7825
        assert result.net == Decimal('7825.00')

    def test_with_allowances(self, engine):
        """Basic + allowances."""
        result = engine.calculate_employee(
            employee_id='E002',
            name='Test',
            basic_salary=Decimal('10000'),
            allowances=Decimal('3000'),
        )

        # Gross: 10000 + 3000 = 13000
        assert result.gross == Decimal('13000.00')
        # Pension: 7% of 10000 = 700 (on basic only)
        assert result.pension_employee == Decimal('700.00')
        # Taxable: 13000 - 700 = 12300
        assert result.taxable == Decimal('12300.00')
        # Tax on 12300:
        #   0-2000: 0
        #   2001-4000: 2000 × 15% = 300
        #   4001-7000: 3000 × 20% = 600
        #   7001-10000: 3000 × 25% = 750
        #   10001-12300: 2300 × 30% = 690
        #   Total = 2340
        assert result.tax == Decimal('2340.00')
        # Net: 13000 - 700 - 2340 = 9960
        assert result.net == Decimal('9960.00')

    def test_zero_salary(self, engine):
        """Zero salary should produce zero everything."""
        result = engine.calculate_employee(
            employee_id='E003',
            name='Zero',
            basic_salary=Decimal('0'),
        )

        assert result.gross == Decimal('0')
        assert result.tax == Decimal('0')
        assert result.net == Decimal('0')
        assert result.pension_employee == Decimal('0')

    def test_low_salary_no_tax(self, engine):
        """Salary below 2000 ETB should have zero tax."""
        result = engine.calculate_employee(
            employee_id='E004',
            name='Low',
            basic_salary=Decimal('1500'),
        )

        # Pension: 7% of 1500 = 105
        assert result.pension_employee == Decimal('105.00')
        # Taxable: 1500 - 105 = 1395 (below 2000 bracket)
        assert result.taxable == Decimal('1395.00')
        # Tax: 0 (below first bracket)
        assert result.tax == Decimal('0.00')
        # Net: 1500 - 105 = 1395
        assert result.net == Decimal('1395.00')

    def test_high_salary_top_bracket(self, engine):
        """High salary should hit the 35% bracket."""
        result = engine.calculate_employee(
            employee_id='E005',
            name='High',
            basic_salary=Decimal('50000'),
        )

        # Pension: 7% of 50000 = 3500
        assert result.pension_employee == Decimal('3500.00')
        # Taxable: 50000 - 3500 = 46500
        assert result.taxable == Decimal('46500.00')
        # Tax on 46500:
        #   0-2000: 0
        #   2001-4000: 2000 × 15% = 300
        #   4001-7000: 3000 × 20% = 600
        #   7001-10000: 3000 × 25% = 750
        #   10001-14000: 4000 × 30% = 1200
        #   14001-46500: 32500 × 35% = 11375
        #   Total = 14225
        assert result.tax == Decimal('14225.00')
        # Net: 50000 - 3500 - 14225 = 32275
        assert result.net == Decimal('32275.00')

    def test_pension_on_basic_only(self, engine):
        """Pension must be calculated on basic salary, not gross."""
        result = engine.calculate_employee(
            employee_id='E006',
            name='Test',
            basic_salary=Decimal('10000'),
            allowances=Decimal('5000'),
        )

        # Pension on basic (10000), not gross (15000)
        assert result.pension_employee == Decimal('700.00')  # 7% of 10000
        assert result.pension_employer == Decimal('1100.00')  # 11% of 10000

    def test_pension_before_tax(self, engine):
        """Pension must be deducted before tax calculation."""
        result = engine.calculate_employee(
            employee_id='E007',
            name='Test',
            basic_salary=Decimal('10000'),
        )

        # Taxable = gross - pension = 10000 - 700 = 9300
        assert result.taxable == Decimal('9300.00')
        # NOT 10000 (which would be wrong)
        assert result.taxable != result.gross

    def test_pension_tax_savings(self, engine):
        """Pension-before-tax should produce measurable savings."""
        result = engine.calculate_employee(
            employee_id='E008',
            name='Test',
            basic_salary=Decimal('10000'),
        )

        # Tax without pension: on 10000
        #   0-2000: 0, 2001-4000: 300, 4001-7000: 600, 7001-10000: 750 = 1650
        # Tax with pension: on 9300
        #   0-2000: 0, 2001-4000: 300, 4001-7000: 600, 7001-9300: 575 = 1475
        # Savings: 1650 - 1475 = 175
        assert result.pension_tax_savings == Decimal('175.00')

    def test_negative_basic_raises(self, engine):
        """Negative basic salary must raise ValueError."""
        with pytest.raises(ValueError, match='cannot be negative'):
            engine.calculate_employee(
                employee_id='E009',
                name='Bad',
                basic_salary=Decimal('-1000'),
            )


# ---------------------------------------------------------------------------
# Test: Calculation flow (explainability)
# ---------------------------------------------------------------------------


class TestCalculationFlow:
    """Every result must have a complete, auditable calculation trail."""

    def test_steps_populated(self, engine):
        """Each employee result must have calculation steps."""
        result = engine.calculate_employee(
            employee_id='E001',
            name='Test',
            basic_salary=Decimal('10000'),
            allowances=Decimal('2000'),
        )

        assert len(result.steps) >= 5  # At minimum: basic, allowances, gross, pension, taxable, tax, net

    def test_steps_have_required_fields(self, engine):
        """Each step must have label, formula, inputs, result."""
        result = engine.calculate_employee(
            employee_id='E001',
            name='Test',
            basic_salary=Decimal('10000'),
        )

        for step in result.steps:
            assert step.label, f'Step {step.step_number} missing label'
            assert step.formula, f'Step {step.step_number} missing formula'
            assert isinstance(step.inputs, dict), f'Step {step.step_number} inputs not a dict'
            assert isinstance(step.result, Decimal), f'Step {step.step_number} result not Decimal'

    def test_steps_sequential(self, engine):
        """Step numbers must be sequential."""
        result = engine.calculate_employee(
            employee_id='E001',
            name='Test',
            basic_salary=Decimal('10000'),
        )

        for i, step in enumerate(result.steps):
            assert step.step_number == i + 1

    def test_legal_references(self, engine):
        """Key steps must have legal references."""
        result = engine.calculate_employee(
            employee_id='E001',
            name='Test',
            basic_salary=Decimal('10000'),
        )

        # Pension step should reference the proclamation
        pension_steps = [s for s in result.steps if 'Pension' in s.label]
        assert any(s.legal_reference for s in pension_steps)

        # Tax step should reference the proclamation
        tax_steps = [s for s in result.steps if 'Income Tax' in s.label]
        assert any(s.legal_reference for s in tax_steps)

    def test_tax_explanation_bilingual(self, engine):
        """Tax explanation should include both Amharic and English."""
        result = engine.calculate_employee(
            employee_id='E001',
            name='Test',
            basic_salary=Decimal('10000'),
        )

        assert result.tax_explanation
        assert 'ETB' in result.tax_explanation  # English
        # Amharic characters present
        assert any(ord(c) > 0x1200 for c in result.tax_explanation)

    def test_tax_breakdown_brackets(self, engine):
        """Tax breakdown must include bracket-by-bracket detail."""
        result = engine.calculate_employee(
            employee_id='E001',
            name='Test',
            basic_salary=Decimal('10000'),
        )

        assert result.tax_breakdown
        assert 'brackets' in result.tax_breakdown
        assert len(result.tax_breakdown['brackets']) > 0

        for bracket in result.tax_breakdown['brackets']:
            assert 'lower' in bracket
            assert 'rate_pct' in bracket
            assert 'taxable_amount' in bracket
            assert 'bracket_tax' in bracket

    def test_effective_tax_rate(self, engine):
        """Effective tax rate must be calculated."""
        result = engine.calculate_employee(
            employee_id='E001',
            name='Test',
            basic_salary=Decimal('10000'),
        )

        assert result.effective_tax_rate > 0
        assert result.effective_tax_rate < 100


# ---------------------------------------------------------------------------
# Test: Exception detection
# ---------------------------------------------------------------------------


class TestExceptions:
    """The engine must detect and report anomalies."""

    def test_duplicate_ids_detected(self, engine):
        """Duplicate employee IDs must be flagged as BLOCK."""
        employees = [
            {'employee_id': 'E001', 'name': 'Abebe', 'basic_salary': 10000, 'bank_account': 'bank:cbe:1000123456789'},
            {'employee_id': 'E001', 'name': 'Abebe', 'basic_salary': 10000, 'bank_account': 'bank:cbe:1000123456789'},
        ]

        result = engine.run_from_data(employees)
        dupes = [e for e in result.exceptions if e.rule_code == 'DUPLICATE_ID']
        assert len(dupes) > 0
        assert all(e.severity == 'BLOCK' for e in dupes)

    def test_missing_bank_detected(self, engine):
        """Missing bank account must be flagged as BLOCK."""
        employees = [
            {'employee_id': 'E001', 'name': 'No Bank', 'basic_salary': 10000, 'bank_account': ''},
        ]

        result = engine.run_from_data(employees)
        missing = [e for e in result.exceptions if e.rule_code == 'MISSING_BANK']
        assert len(missing) == 1
        assert missing[0].severity == 'BLOCK'

    def test_negative_net_detected(self, engine):
        """Negative net pay must be flagged as BLOCK."""
        # This shouldn't normally happen with valid inputs, but the guard exists
        employees = [
            {'employee_id': 'E001', 'name': 'Test', 'basic_salary': 100, 'bank_account': 'bank:cbe:1000123456789'},
        ]

        result = engine.run_from_data(employees)
        # With 100 ETB basic, pension = 7, taxable = 93, tax = 0, net = 93
        # No negative net expected here, but let's verify the guard works
        negatives = [e for e in result.exceptions if e.rule_code == 'NEGATIVE_NET']
        assert len(negatives) == 0  # Should be clean

    def test_high_salary_flagged(self, engine):
        """Salary > 500k ETB must be flagged."""
        employees = [
            {'employee_id': 'E001', 'name': 'Rich', 'basic_salary': 600000, 'bank_account': 'bank:cbe:1000123456789'},
        ]

        result = engine.run_from_data(employees)
        high = [e for e in result.exceptions if e.rule_code == 'SALARY_HIGH']
        assert len(high) == 1
        assert high[0].severity == 'FLAG'

    def test_missing_tin_warned(self, engine):
        """Missing TIN must be flagged as WARN."""
        employees = [
            {'employee_id': 'E001', 'name': 'No TIN', 'basic_salary': 10000, 'bank_account': 'bank:cbe:1000123456789', 'tin': ''},
        ]

        result = engine.run_from_data(employees)
        missing_tin = [e for e in result.exceptions if e.rule_code == 'MISSING_TIN']
        assert len(missing_tin) == 1
        assert missing_tin[0].severity == 'WARN'

    def test_salary_change_detected(self, engine):
        """Salary change > 30% must be flagged."""
        employees = [
            {'employee_id': 'E001', 'name': 'Test', 'basic_salary': 20000, 'bank_account': 'bank:cbe:1000123456789'},
        ]

        previous = {
            'E001': {'basic': 10000, 'allowances': 0, 'name': 'Test'},
        }

        result = engine.run_from_data(employees, previous_payslips=previous)
        changes = [e for e in result.exceptions if e.rule_code == 'SALARY_CHANGE_30PCT']
        assert len(changes) == 1
        assert changes[0].severity == 'FLAG'

    def test_cash_compliance_detected(self, engine):
        """Net > 50k without bank must be flagged."""
        employees = [
            {'employee_id': 'E001', 'name': 'Cash', 'basic_salary': 80000, 'bank_account': ''},
        ]

        result = engine.run_from_data(employees)
        cash = [e for e in result.exceptions if e.rule_code == 'CASH_COMPLIANCE']
        assert len(cash) == 1

    def test_clean_data_no_blocks(self, engine, sample_employees):
        """Clean data should produce no BLOCK exceptions."""
        result = engine.run_from_data(sample_employees)
        assert result.block_count == 0

    def test_exception_counts(self, engine):
        """Exception counts must be accurate."""
        employees = [
            {'employee_id': 'E001', 'name': 'OK', 'basic_salary': 10000, 'bank_account': 'bank:cbe:1000123456789', 'tin': '123'},
            {'employee_id': 'E002', 'name': 'No Bank', 'basic_salary': 10000, 'bank_account': '', 'tin': '456'},
            {'employee_id': 'E003', 'name': 'No TIN', 'basic_salary': 10000, 'bank_account': 'bank:cbe:1000987654321', 'tin': ''},
        ]

        result = engine.run_from_data(employees)
        assert result.block_count >= 1  # Missing bank
        assert result.warn_count >= 1  # Missing TIN


# ---------------------------------------------------------------------------
# Test: Change detection
# ---------------------------------------------------------------------------


class TestChangeDetection:
    """Changes vs previous period must be detected."""

    def test_new_hires(self, engine):
        """New employees must be detected."""
        employees = [
            {'employee_id': 'E001', 'name': 'Existing', 'basic_salary': 10000, 'bank_account': 'bank:cbe:1000123456789'},
            {'employee_id': 'E002', 'name': 'New Hire', 'basic_salary': 8000, 'bank_account': 'bank:cbe:1000987654321'},
        ]

        previous = {
            'E001': {'basic': 10000, 'allowances': 0, 'name': 'Existing'},
        }

        result = engine.run_from_data(employees, previous_payslips=previous)
        assert len(result.new_hires) == 1
        assert result.new_hires[0]['employee_id'] == 'E002'

    def test_departures(self, engine):
        """Departing employees must be detected."""
        employees = [
            {'employee_id': 'E001', 'name': 'Staying', 'basic_salary': 10000, 'bank_account': 'bank:cbe:1000123456789'},
        ]

        previous = {
            'E001': {'basic': 10000, 'allowances': 0, 'name': 'Staying'},
            'E002': {'basic': 8000, 'allowances': 0, 'name': 'Leaving'},
        }

        result = engine.run_from_data(employees, previous_payslips=previous)
        assert len(result.departures) == 1
        assert result.departures[0]['employee_id'] == 'E002'

    def test_salary_changes(self, engine):
        """Salary changes must be detected."""
        employees = [
            {'employee_id': 'E001', 'name': 'Test', 'basic_salary': 15000, 'bank_account': 'bank:cbe:1000123456789'},
        ]

        previous = {
            'E001': {'basic': 10000, 'allowances': 0, 'name': 'Test'},
        }

        result = engine.run_from_data(employees, previous_payslips=previous)
        assert len(result.salary_changes) == 1
        assert result.salary_changes[0]['delta'] == Decimal('5000')

    def test_gross_delta(self, engine):
        """Gross delta must be calculated."""
        employees = [
            {'employee_id': 'E001', 'name': 'Test', 'basic_salary': 15000, 'bank_account': 'bank:cbe:1000123456789'},
        ]

        previous = {
            'E001': {'basic': 10000, 'allowances': 0, 'name': 'Test'},
        }

        result = engine.run_from_data(employees, previous_payslips=previous)
        assert result.gross_delta == Decimal('5000')
        assert result.gross_delta_pct == Decimal('50.0')


# ---------------------------------------------------------------------------
# Test: Approval workflow
# ---------------------------------------------------------------------------


class TestApprovalWorkflow:
    """Approval must follow the state machine."""

    def test_approve_clean_run(self, engine, sample_employees):
        """Clean run can be approved."""
        result = engine.run_from_data(sample_employees)
        result = engine.approve(result, approved_by='Manager', notes='Looks good')

        assert result.status == 'approved'
        assert result.approved_by == 'Manager'
        assert result.approved_at is not None

    def test_cannot_approve_with_blocks(self, engine):
        """Cannot approve if BLOCK exceptions exist."""
        employees = [
            {'employee_id': 'E001', 'name': 'No Bank', 'basic_salary': 10000, 'bank_account': ''},
        ]

        result = engine.run_from_data(employees)
        with pytest.raises(ValueError, match='BLOCK'):
            engine.approve(result, approved_by='Manager')

    def test_override_flag(self, engine):
        """FLAG exceptions can be overridden."""
        employees = [
            {'employee_id': 'E001', 'name': 'Rich', 'basic_salary': 600000, 'bank_account': 'bank:cbe:1000123456789'},
        ]

        result = engine.run_from_data(employees)
        assert result.flag_count > 0

        # Override the flag
        for exc in result.exceptions:
            if exc.severity == 'FLAG':
                result = engine.override_exception(
                    result,
                    employee_id=exc.employee_id,
                    rule_code=exc.rule_code,
                    reason='Executive salary confirmed by board',
                    overridden_by='CEO',
                )

        assert result.flag_count == 0

    def test_cannot_override_block(self, engine):
        """BLOCK exceptions cannot be overridden."""
        employees = [
            {'employee_id': 'E001', 'name': 'No Bank', 'basic_salary': 10000, 'bank_account': ''},
        ]

        result = engine.run_from_data(employees)
        block = [e for e in result.exceptions if e.severity == 'BLOCK'][0]

        with pytest.raises(ValueError, match='BLOCK'):
            engine.override_exception(
                result,
                employee_id=block.employee_id,
                rule_code=block.rule_code,
                reason='Trying to override',
                overridden_by='Manager',
            )

    def test_lock_approved_run(self, engine, sample_employees):
        """Approved run can be locked."""
        result = engine.run_from_data(sample_employees)
        result = engine.approve(result, approved_by='Manager')
        result = engine.lock(result, locked_by='Finance')

        assert result.status == 'locked'

    def test_cannot_lock_draft(self, engine, sample_employees):
        """Draft run cannot be locked."""
        result = engine.run_from_data(sample_employees)
        with pytest.raises(ValueError, match='Cannot lock'):
            engine.lock(result, locked_by='Finance')


# ---------------------------------------------------------------------------
# Test: Bank file generation
# ---------------------------------------------------------------------------


class TestBankFile:
    """Bank file must be generated for clean runs."""

    def test_bank_file_generated(self, engine, sample_employees):
        """Clean run should generate a bank file."""
        result = engine.run_from_data(sample_employees)

        assert result.bank_file_data
        assert result.bank_file_name

    def test_bank_file_not_generated_with_blocks(self, engine):
        """Bank file should NOT be generated if BLOCK exceptions exist."""
        employees = [
            {'employee_id': 'E001', 'name': 'No Bank', 'basic_salary': 10000, 'bank_account': ''},
        ]

        result = engine.run_from_data(employees)
        assert not result.bank_file_data

    def test_bank_file_csv_format(self, engine, sample_employees):
        """Bank file must be valid CSV."""
        result = engine.run_from_data(sample_employees)

        import csv
        import io

        reader = csv.reader(io.StringIO(result.bank_file_data.decode('utf-8')))
        rows = list(reader)

        # Header + 3 employees
        assert len(rows) == 4
        assert rows[0] == ['account_number', 'amount', 'narrative', 'currency']

    def test_bank_file_amounts_match_net(self, engine, sample_employees):
        """Bank file amounts must match net pay."""
        result = engine.run_from_data(sample_employees)

        import csv
        import io

        reader = csv.reader(io.StringIO(result.bank_file_data.decode('utf-8')))
        next(reader)  # skip header

        bank_amounts = []
        for row in reader:
            bank_amounts.append(Decimal(row[1]))

        net_amounts = [e.net for e in result.employees]
        assert sorted(bank_amounts) == sorted(net_amounts)


# ---------------------------------------------------------------------------
# Test: Excel export
# ---------------------------------------------------------------------------


class TestExcelExport:
    """Excel export must produce valid workbooks."""

    def test_export_produces_bytes(self, engine, sample_employees):
        """Export must return bytes."""
        result = engine.run_from_data(sample_employees)
        xlsx = engine.export_to_excel(result)

        assert isinstance(xlsx, bytes)
        assert len(xlsx) > 0

    def test_export_valid_xlsx(self, engine, sample_employees):
        """Export must be a valid XLSX file."""
        import openpyxl

        result = engine.run_from_data(sample_employees)
        xlsx = engine.export_to_excel(result)

        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        assert wb.sheetnames  # Has sheets
        wb.close()

    def test_export_has_all_sheets(self, engine, sample_employees):
        """Export must have all required sheets."""
        import openpyxl

        result = engine.run_from_data(sample_employees)
        xlsx = engine.export_to_excel(result)

        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        expected_sheets = ['Summary', 'Payroll', 'Calculation Flow', 'Tax Breakdown', 'Exceptions', 'Changes', 'Bank File', 'Approval']
        for sheet in expected_sheets:
            assert sheet in wb.sheetnames, f'Missing sheet: {sheet}'
        wb.close()

    def test_export_summary_has_hashes(self, engine, sample_employees):
        """Summary sheet must include determinism hashes."""
        import openpyxl

        result = engine.run_from_data(sample_employees)
        xlsx = engine.export_to_excel(result)

        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws = wb['Summary']

        # Find the hash rows
        found_input_hash = False
        found_calc_hash = False
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0] == 'Input Hash':
                found_input_hash = True
            if row[0] == 'Calculation Hash':
                found_calc_hash = True

        assert found_input_hash, 'Input Hash not found in Summary'
        assert found_calc_hash, 'Calculation Hash not found in Summary'
        wb.close()

    def test_export_payroll_row_count(self, engine, sample_employees):
        """Payroll sheet must have one row per employee + header + totals."""
        import openpyxl

        result = engine.run_from_data(sample_employees)
        xlsx = engine.export_to_excel(result)

        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws = wb['Payroll']

        # Header (row 1) + 3 employees (rows 2-4) + totals (row 5)
        assert ws.max_row == 5
        wb.close()


# ---------------------------------------------------------------------------
# Test: Edge cases
# ---------------------------------------------------------------------------


class TestEdgeCases:
    """Handle unusual inputs gracefully."""

    def test_empty_employee_list(self, engine):
        """Empty employee list should produce a valid (empty) result."""
        result = engine.run_from_data([])

        assert result.employee_count == 0
        assert result.total_gross == Decimal('0')
        assert result.total_net == Decimal('0')

    def test_single_employee(self, engine):
        """Single employee should work fine."""
        employees = [
            {'employee_id': 'E001', 'name': 'Solo', 'basic_salary': 10000, 'bank_account': 'bank:cbe:1000123456789'},
        ]

        result = engine.run_from_data(employees)
        assert result.employee_count == 1
        assert result.total_gross == result.employees[0].gross

    def test_string_salary_parsed(self, engine):
        """String salary values should be parsed correctly."""
        employees = [
            {'employee_id': 'E001', 'name': 'Test', 'basic_salary': '10000', 'allowances': '2000', 'bank_account': 'bank:cbe:1000123456789'},
        ]

        result = engine.run_from_data(employees)
        assert result.employees[0].basic_salary == Decimal('10000')
        assert result.employees[0].gross == Decimal('12000')

    def test_zero_allowances(self, engine):
        """Zero allowances should not cause errors."""
        result = engine.calculate_employee(
            employee_id='E001',
            name='Test',
            basic_salary=Decimal('10000'),
            allowances=Decimal('0'),
        )

        assert result.gross == Decimal('10000')
        assert result.taxable_allowances == Decimal('0')

    def test_very_small_salary(self, engine):
        """Very small salary (1 ETB) should work."""
        result = engine.calculate_employee(
            employee_id='E001',
            name='Test',
            basic_salary=Decimal('1'),
        )

        assert result.gross == Decimal('1')
        assert result.pension_employee == Decimal('0.07')
        assert result.tax == Decimal('0')
        assert result.net == Decimal('0.93')

    def test_batch_vs_individual(self, engine):
        """Batch processing must produce same results as individual calculation."""
        employees = [
            {'employee_id': 'E001', 'name': 'A', 'basic_salary': 10000, 'bank_account': 'bank:cbe:1000123456789'},
            {'employee_id': 'E002', 'name': 'B', 'basic_salary': 20000, 'bank_account': 'bank:cbe:1000987654321'},
        ]

        batch_result = engine.run_from_data(employees)

        individual_a = engine.calculate_employee('E001', 'A', Decimal('10000'))
        individual_b = engine.calculate_employee('E002', 'B', Decimal('20000'))

        assert batch_result.employees[0].gross == individual_a.gross
        assert batch_result.employees[0].tax == individual_a.tax
        assert batch_result.employees[0].net == individual_a.net
        assert batch_result.employees[1].gross == individual_b.gross
        assert batch_result.employees[1].tax == individual_b.tax
        assert batch_result.employees[1].net == individual_b.net


# ---------------------------------------------------------------------------
# Test: Import from Excel
# ---------------------------------------------------------------------------


class TestExcelImport:
    """Import from Excel/CSV files."""

    def test_import_from_csv(self, engine):
        """Import from CSV file."""
        import csv
        import tempfile

        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
            writer = csv.writer(f)
            writer.writerow(['employee_id', 'name', 'basic_salary', 'allowances', 'bank_account'])
            writer.writerow(['E001', 'Abebe', 10000, 2000, 'bank:cbe:1000123456789'])
            writer.writerow(['E002', 'Tigist', 15000, 3000, 'bank:cbe:1000987654321'])
            filepath = f.name

        try:
            result = engine.run_from_excel(filepath)
            assert result.employee_count == 2
            assert result.total_gross > 0
        finally:
            os.unlink(filepath)

    def test_import_from_xlsx(self, engine):
        """Import from Excel file."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['employee_id', 'name', 'basic_salary', 'allowances', 'bank_account'])
        ws.append(['E001', 'Abebe', 10000, 2000, 'bank:cbe:1000123456789'])
        ws.append(['E002', 'Tigist', 15000, 3000, 'bank:cbe:1000987654321'])

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            filepath = f.name

        try:
            result = engine.run_from_excel(filepath)
            assert result.employee_count == 2
            assert result.total_gross > 0
        finally:
            os.unlink(filepath)
